"""Deterministic qa_agent FAQ table contracts shared by migration and Runner."""
from __future__ import annotations

import csv
import hashlib
import io
import re
from pathlib import Path
from typing import Any, Iterable


FAQ_TYPE_CODE = "qa-agent-faq"
FAQ_PROFILE_CODE = "qa-agent-faq-default"
FAQ_COLLECTION_NAME = "dataforge_qa_agent_faq"
FAQ_TEMPLATE_CODE = "qa-agent-faq-structured"
FAQ_FILENAME_PATTERN = re.compile(r"^faq-([A-Za-z0-9_-]+)\.(csv|xlsx)$", re.IGNORECASE)

FAQ_SCHEMA: dict[str, Any] = {
    "type": "object",
    "required": ["aq_id", "org_code", "question_name", "answer_desc", "full_text"],
    "properties": {
        "aq_id": {"type": "string", "minLength": 1},
        "org_code": {"type": "string", "minLength": 1},
        "question_name": {"type": "string", "minLength": 1},
        "answer_desc": {"type": "string", "minLength": 1},
        "full_text": {"type": "string", "minLength": 1},
        "doc_id": {"type": "string"},
        "document_id": {"type": "string"},
        "ref_doc_id": {"type": "string"},
    },
}

_HEADER_ALIASES = {
    "aq_id": {"aq_id", "id", "faq_id", "faq id", "faq编号", "faq 编号"},
    "org_code": {"org_code", "org code", "机构编码", "机构代码"},
    "question_name": {"question_name", "question", "问题", "问题名称"},
    "answer_desc": {"answer_desc", "answer", "答案", "回答"},
    "doc_id": {"doc_id", "doc id"},
    "document_id": {"document_id", "document id"},
    "ref_doc_id": {"ref_doc_id", "ref doc id"},
}


def clean_faq_text(value: Any) -> str:
    """Normalize table cells without changing FAQ business wording."""
    text = str("" if value is None else value).replace("\ufeff", "")
    return re.sub(r"\s+", " ", text).strip()


def faq_org_code_from_filename(filename: str) -> str:
    """Return the institution encoded by ``faq-{org_code}.csv|xlsx``."""
    basename = Path(str(filename or "")).name
    match = FAQ_FILENAME_PATTERN.fullmatch(basename)
    if not match:
        raise ValueError("FAQ 文件名必须为 faq-{org_code}.csv 或 faq-{org_code}.xlsx")
    return match.group(1)


def _header_key(value: Any) -> str:
    return clean_faq_text(value).casefold()


_CANONICAL_HEADERS = {
    _header_key(alias): canonical
    for canonical, aliases in _HEADER_ALIASES.items()
    for alias in aliases
}


def _rows_from_csv(payload: bytes) -> list[dict[str, Any]]:
    text = payload.decode("utf-8-sig", errors="strict")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("FAQ CSV 缺少表头")
    return [
        {"row_number": index, "sheet": "", "values": dict(row)}
        for index, row in enumerate(reader, start=2)
        if any(clean_faq_text(value) for value in row.values())
    ]


def _rows_from_xlsx(payload: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    output: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers: list[str] | None = None
        header_row = 0
        for row_number, row in enumerate(rows, start=1):
            if not any(clean_faq_text(value) for value in row):
                continue
            if headers is None:
                headers = [clean_faq_text(value) for value in row]
                header_row = row_number
                continue
            values = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            if any(clean_faq_text(value) for value in values.values()):
                output.append({"row_number": row_number, "sheet": sheet.title, "values": values})
        if headers is None:
            continue
        if not any(headers):
            raise ValueError(f"FAQ XLSX 工作表 {sheet.title} 第 {header_row} 行表头为空")
    return output


def parse_table_rows(filename: str, payload: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv(payload)
    if suffix == ".xlsx":
        return _rows_from_xlsx(payload)
    raise ValueError("qa-agent-faq 模板只支持 CSV 或 XLSX")


def normalize_faq_rows(documents: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    documents = list(documents)
    if len(documents) != 1:
        raise ValueError("qa-agent-faq 文档库每次只能处理一个权威文件")
    document = documents[0]
    filename = str(document.get("filename") or "")
    filename_org_code = faq_org_code_from_filename(filename)
    raw_rows = list(document.get("table_rows") or [])
    if not raw_rows:
        raise ValueError("FAQ 文件没有可处理的数据行")
    normalized: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for raw in raw_rows:
        values = dict(raw.get("values") or {})
        mapped: dict[str, str] = {}
        org_header_present = False
        for header, value in values.items():
            canonical = _CANONICAL_HEADERS.get(_header_key(header))
            if not canonical:
                continue
            if canonical == "org_code":
                org_header_present = True
            cleaned = clean_faq_text(value)
            if canonical in mapped and mapped[canonical] != cleaned:
                raise ValueError(f"FAQ 第 {raw.get('row_number')} 行字段 {canonical} 使用多个别名且值不一致")
            mapped[canonical] = cleaned
        row_org_code = mapped.get("org_code")
        if org_header_present and row_org_code != filename_org_code:
            location = f"{raw.get('sheet') or filename}:{raw.get('row_number')}"
            raise ValueError(
                f"FAQ {location} 的 org_code={row_org_code} 与文件名机构 {filename_org_code} 不一致"
            )
        mapped["org_code"] = filename_org_code
        missing = [key for key in ("aq_id", "question_name", "answer_desc") if not mapped.get(key)]
        if missing:
            location = f"{raw.get('sheet') or document.get('filename')}:{raw.get('row_number')}"
            raise ValueError(f"FAQ {location} 缺少必填字段：{', '.join(missing)}")
        aq_id = mapped["aq_id"]
        if aq_id in seen_ids:
            raise ValueError(f"FAQ aq_id 重复：{aq_id}")
        seen_ids.add(aq_id)
        mapped["full_text"] = f"{mapped['question_name']}: {mapped['answer_desc']}"
        normalized.append({
            **mapped,
            "row_number": int(raw.get("row_number") or len(normalized) + 2),
            "sheet": clean_faq_text(raw.get("sheet")),
        })
    return normalized


def faq_rows_digest(rows: Iterable[dict[str, Any]]) -> str:
    digest = hashlib.sha256()
    for row in sorted(rows, key=lambda item: (str(item.get("org_code")), str(item.get("aq_id")))):
        line = "\x1f".join(str(row.get(key) or "") for key in (
            "aq_id", "org_code", "question_name", "answer_desc", "doc_id", "document_id", "ref_doc_id",
        ))
        digest.update(line.encode("utf-8")); digest.update(b"\n")
    return digest.hexdigest()


def faq_template_definition() -> dict[str, Any]:
    nodes = [
        {"id": "reviewed-input", "kind": "operator", "ref": "reviewed-source-chunk-input"},
        {"id": "generate", "kind": "operator", "ref": "faq-record-mapper", "params": {"knowledge_type": FAQ_TYPE_CODE}},
        {"id": "evaluate", "kind": "operator", "ref": "quality-evaluator", "params": {"knowledge_type": FAQ_TYPE_CODE, "quality_profile_revision_id": "qualityrev_default"}},
        {"id": "filter", "kind": "operator", "ref": "quality-filter", "params": {"knowledge_type": FAQ_TYPE_CODE, "quality_profile_revision_id": "qualityrev_default"}},
        {"id": "bind", "kind": "operator", "ref": "source-binding", "params": {"knowledge_type": FAQ_TYPE_CODE}},
        {"id": "validate", "kind": "operator", "ref": "schema-validator", "params": {"knowledge_type": FAQ_TYPE_CODE}},
        {"id": "diff", "kind": "operator", "ref": "knowledge-diff", "params": {"knowledge_type": FAQ_TYPE_CODE}},
        {"id": "sink", "kind": "knowledge_sink", "knowledge_type": FAQ_TYPE_CODE, "output_key": FAQ_TYPE_CODE},
    ]
    order = ["reviewed-input", "generate", "evaluate", "filter", "bind", "validate", "diff", "sink"]
    return {
        "schema_version": 3,
        "purpose": "knowledge",
        "nodes": nodes,
        "edges": [
            {"source": source, "source_port": "output", "target": target, "target_port": "input"}
            for source, target in zip(order, order[1:])
        ],
        "ui": {"positions": {}},
    }

"""Fixed V7 knowledge processing runner.

The runner has no candidate-confirm step: each successful job atomically writes
the current state and change history of the explicitly selected V7 library.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import re
import tempfile
import threading
import time
import unicodedata
from copy import deepcopy
from pathlib import Path
from typing import Any

import httpx
from fastapi import FastAPI, Header, HTTPException
from openai import APITimeoutError
from pydantic import BaseModel
from sqlalchemy import select

from ..config import Settings
from .catalog import catalog_by_code, normalize_chunker_params
from .operators import OperatorExecutionContext, build_builtin_registry
from .graph_literal import classify_object, detect_literal
from .graph_prompt import entity_prompt_for, relation_prompt_for
from .graph_quality import evaluate_graph_quality
from .graph_schema import GraphExtractionConfig, normalize_graph_config
from .llm_serving import DEFAULT_LLM_SERVING_ID, configure_llm_serving_registry, get_llm_serving_registry
from .models import KnowledgeJob, Source, SourceVersion
from .parser_runtime import content_list_blocks, parse_with_mineru
from .source_anchor import finalize_source_blocks, sort_positions
from .storage import LocalObjectStore, MinioObjectStore
from .store import V7Store
from .faq import FAQ_FILENAME_PATTERN, FAQ_TYPE_CODE, normalize_faq_rows, parse_table_rows


logger = logging.getLogger("dataforge.v7.runner")


class RunRequest(BaseModel):
    job_id: str | None = None
    source_preparation_job_id: str | None = None
    flow_run_id: str | None = None
    lease_owner: str | None = None


def _diagnostic_pdf() -> bytes:
    """Build a valid one-page PDF without adding a runtime PDF dependency."""
    stream = b"BT /F1 12 Tf 20 100 Td (DataForge health) Tj ET"
    objects = [
        b"<</Type/Catalog/Pages 2 0 R>>",
        b"<</Type/Pages/Kids[3 0 R]/Count 1>>",
        b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 200 200]/Resources<</Font<</F1 4 0 R>>>>/Contents 5 0 R>>",
        b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
        b"<</Length %d>>stream\n" % len(stream) + stream + b"\nendstream",
    ]
    payload = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, body in enumerate(objects, 1):
        offsets.append(len(payload)); payload.extend(f"{index} 0 obj\n".encode()); payload.extend(body); payload.extend(b"\nendobj\n")
    xref = len(payload); payload.extend(f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n".encode())
    for offset in offsets[1:]: payload.extend(f"{offset:010d} 00000 n \n".encode())
    payload.extend(f"trailer<</Root 1 0 R/Size {len(objects)+1}>>\nstartxref\n{xref}\n%%EOF\n".encode())
    return bytes(payload)


def _objects(settings: Settings):
    if settings.minio_endpoint and settings.minio_access_key and settings.minio_secret_key:
        return MinioObjectStore(settings.minio_endpoint, settings.minio_access_key, settings.minio_secret_key, settings.minio_bucket)
    return LocalObjectStore(settings.state_dir / "v7-objects")


def _extract_text(filename: str, payload: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md"}: return payload.decode("utf-8-sig", errors="replace")
    if suffix == ".csv":
        rows = csv.reader(io.StringIO(payload.decode("utf-8-sig", errors="replace")))
        return "\n".join(" | ".join(cell.strip() for cell in row if cell.strip()) for row in rows)
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"[Sheet: {sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    lines.append(" | ".join(values))
        return "\n".join(lines)
    if suffix == ".docx":
        text, _ = _docx_source_blocks(payload)
        return text
    if suffix == ".doc":
        with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as handle:
            handle.write(payload); name = handle.name
        try:
            import subprocess
            return subprocess.run(["antiword", name], capture_output=True, check=True, text=True, encoding="utf-8", errors="replace").stdout
        finally:
            Path(name).unlink(missing_ok=True)
    raise ValueError("不支持的文档类型")


def _docx_source_blocks(payload: bytes) -> tuple[str, list[dict[str, Any]]]:
    """Read DOCX paragraphs and table rows in their original document order."""
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    document = Document(io.BytesIO(payload))
    raw_blocks: list[dict[str, Any]] = []
    table_index = 0
    children = document.iter_inner_content() if hasattr(document, "iter_inner_content") else document.paragraphs
    for child in children:
        if isinstance(child, Paragraph):
            text = child.text.strip()
            if not text:
                continue
            style = str(getattr(getattr(child, "style", None), "name", "") or "")
            heading = re.search(r"(?:Heading|标题)\s*([1-9])", style, re.I)
            block_type = "heading" if heading else "paragraph"
            raw_blocks.append({
                "block_id": f"docx:{len(raw_blocks)}",
                "block_type": block_type,
                "heading_level": int(heading.group(1)) if heading else None,
                "style": style,
                "text": text,
            })
        elif isinstance(child, Table):
            for row_index, row in enumerate(child.rows):
                cells = [cell.text.strip() for cell in row.cells]
                if not any(cells):
                    continue
                raw_blocks.append({
                    "block_id": f"docx:{len(raw_blocks)}",
                    "block_type": "table_row",
                    "table_index": table_index,
                    "row_index": row_index,
                    "cells": cells,
                    "text": " | ".join(cells),
                })
            table_index += 1
    return finalize_source_blocks(raw_blocks)


def _clean_document_text(text: str, ref: str) -> str:
    value = text.replace("\r\n", "\n").replace("\r", "\n")
    if ref == "text-cleaner":
        value = "".join(char for char in value if char in {"\n", "\t"} or unicodedata.category(char) != "Cc")
    elif ref == "whitespace-cleaner":
        value = "\n".join(re.sub(r"[\t \u00a0]+", " ", line).strip() for line in value.split("\n"))
        value = re.sub(r"\n{3,}", "\n\n", value)
    elif ref == "text-normalizer":
        value = unicodedata.normalize("NFKC", value)
    return value.strip()


def _boundary_units(text: str, delimiters: list[str]) -> list[dict[str, Any]]:
    pattern = "(" + "|".join(re.escape(value) for value in sorted(delimiters, key=len, reverse=True)) + ")"
    parts = re.split(pattern, text)
    units: list[dict[str, Any]] = []
    cursor = 0
    for index in range(0, len(parts), 2):
        body = parts[index]
        delimiter = parts[index + 1] if index + 1 < len(parts) else ""
        value = body + delimiter
        if value:
            units.append({"text": value, "start": cursor, "end": cursor + len(value)})
            cursor += len(value)
    return units


def split_document_text(text: str, raw_params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    params = normalize_chunker_params(raw_params)
    target = params["chunk_size"]
    minimum = params["min_chunk_size"]
    units = _boundary_units(text, params["delimiters"])
    expanded: list[dict[str, Any]] = []
    for unit in units:
        value = unit["text"]
        if len(value) <= target:
            expanded.append({**unit, "hard_cut": False})
            continue
        for offset in range(0, len(value), target):
            piece = value[offset:offset + target]
            expanded.append({"text": piece, "start": unit["start"] + offset,
                             "end": unit["start"] + offset + len(piece), "hard_cut": True})

    chunks: list[dict[str, Any]] = []
    current: list[dict[str, Any]] = []
    current_size = 0
    overlap_target = round(target * params["overlap_percent"] / 100)

    def flush() -> None:
        nonlocal current, current_size
        if not current:
            return
        body = "".join(item["text"] for item in current).strip()
        if body:
            heading = next((line.strip() for line in body.splitlines()
                            if re.match(r"^#{1,6}\s+\S", line.strip())), "")
            chunks.append({
                "content": body,
                "char_start": current[0]["start"],
                "char_end": current[-1]["end"],
                "hard_cut": any(item.get("hard_cut") for item in current),
                "heading_context": heading if params["include_heading"] else "",
                "overlap_chars": max(0, sum(len(item["text"]) for item in current if item.get("overlap"))),
            })
        overlap: list[dict[str, Any]] = []
        overlap_size = 0
        if overlap_target:
            for item in reversed(current):
                if overlap_size + len(item["text"]) > overlap_target:
                    break
                overlap.insert(0, {**item, "overlap": True})
                overlap_size += len(item["text"])
                if overlap_size >= overlap_target:
                    break
        current = overlap
        current_size = overlap_size

    for unit in expanded:
        if current and current_size + len(unit["text"]) > target and current_size >= minimum:
            flush()
        current.append(unit)
        current_size += len(unit["text"])
        if unit.get("hard_cut") and current_size >= target:
            flush()
    flush()
    if len(chunks) > 1 and len(chunks[-1]["content"]) < minimum:
        tail = chunks.pop()
        previous = chunks[-1]
        previous["content"] = (previous["content"] + "\n" + tail["content"]).strip()
        previous["char_end"] = tail["char_end"]
        previous["hard_cut"] = previous["hard_cut"] or tail["hard_cut"]
    return chunks


def _page_segments_from_blocks(blocks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[int, list[dict[str, Any]]] = {}
    for block in blocks:
        page_index = block.get("page_index")
        if isinstance(page_index, int):
            grouped.setdefault(page_index, []).append(block)
    return [
        {"page_index": page_index, "page": page_index + 1,
         "text": "\n\n".join(str(block.get("text") or "") for block in page_blocks),
         "source_blocks": page_blocks}
        for page_index, page_blocks in sorted(grouped.items())
    ]


def _clean_source_blocks(blocks: list[dict[str, Any]], ref: str) -> tuple[str, list[dict[str, Any]]]:
    cleaned = [{**block, "text": _clean_document_text(str(block.get("text") or ""), ref)} for block in blocks]
    return finalize_source_blocks(cleaned)


def split_document_blocks(document: dict[str, Any], raw_params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    """Split structured parser blocks while carrying SourceAnchorV2 positions."""
    params = normalize_chunker_params(raw_params)
    blocks = [dict(block) for block in document.get("source_blocks") or [] if str(block.get("text") or "").strip()]
    if not blocks:
        return split_document_text(str(document.get("text") or ""), params)
    groups: list[list[dict[str, Any]]] = []
    if params["preserve_page_boundary"] and document.get("source_type") == "pdf":
        for block in blocks:
            if not groups or groups[-1][0].get("page_index") != block.get("page_index"):
                groups.append([])
            groups[-1].append(block)
    else:
        groups = [blocks]

    results: list[dict[str, Any]] = []
    for group in groups:
        group_parts: list[str] = []
        intervals: list[tuple[int, int, dict[str, Any]]] = []
        cursor = 0
        for block in group:
            if group_parts:
                group_parts.append("\n\n")
                cursor += 2
            text = str(block["text"])
            intervals.append((cursor, cursor + len(text), block))
            group_parts.append(text)
            cursor += len(text)
        group_text = "".join(group_parts)
        for chunk in split_document_text(group_text, params):
            raw_start, raw_end = int(chunk["char_start"]), int(chunk["char_end"])
            raw_value = group_text[raw_start:raw_end]
            offset = raw_value.find(str(chunk["content"]))
            effective_start = raw_start + (offset if offset >= 0 else len(raw_value) - len(raw_value.lstrip()))
            effective_end = effective_start + len(str(chunk["content"]))
            positions: list[dict[str, Any]] = []
            document_ranges: list[tuple[int, int]] = []
            missing_pdf_bbox = False
            for block_start, block_end, block in intervals:
                intersection_start = max(block_start, effective_start)
                intersection_end = min(block_end, effective_end)
                if intersection_end <= intersection_start:
                    continue
                chunk_range = [intersection_start - effective_start, intersection_end - effective_start]
                global_start = int(block.get("char_start", 0)) + intersection_start - block_start
                global_end = int(block.get("char_start", 0)) + intersection_end - block_start
                document_ranges.append((global_start, global_end))
                if document.get("source_type") == "pdf":
                    position = {
                        "kind": "pdf_bbox" if block.get("bbox") else "pdf_page",
                        "block_id": block["block_id"], "block_index": block.get("block_index"),
                        "page": block.get("page"), "page_index": block.get("page_index"),
                        "chunk_range": chunk_range,
                    }
                    if block.get("bbox"):
                        position["bbox"] = list(block["bbox"])
                    else:
                        missing_pdf_bbox = True
                    positions.append(position)
                elif document.get("source_type") == "docx":
                    positions.append({
                        "kind": "docx_block", "block_id": block["block_id"],
                        "block_type": block.get("block_type", "paragraph"),
                        "block_index": block.get("block_index"), "chunk_range": chunk_range,
                    })
            positions = sort_positions(positions)
            pages = [int(position["page"]) for position in positions if isinstance(position.get("page"), int)]
            precision = "block" if positions and any(position.get("kind") != "pdf_page" for position in positions) else (
                "page" if positions else "unavailable"
            )
            anchor: dict[str, Any] = {
                "anchor_version": 2,
                "source_version_id": document.get("source_version_id"),
                "source_type": document.get("source_type"),
                "precision": precision,
                "positions": positions,
                **{key: value for key, value in chunk.items() if key != "content"},
            }
            if document_ranges:
                anchor.update(char_start=min(value[0] for value in document_ranges),
                              char_end=max(value[1] for value in document_ranges))
            if pages:
                anchor.update(page=min(pages), page_index=min(pages) - 1,
                              page_start=min(pages), page_end=max(pages))
            if missing_pdf_bbox and any(position.get("kind") == "pdf_bbox" for position in positions):
                anchor["position_status"] = "partial"
            results.append({"content": chunk["content"], "anchor": anchor})
    return results


def _chunks(text: str, size: int = 800) -> list[str]:
    return [item["content"] for item in split_document_text(text, {"chunk_size": size})]


def _operator_params(definition: dict[str, Any] | None, ref: str) -> dict[str, Any]:
    node = next((item for item in (definition or {}).get("nodes", []) if item.get("ref") == ref), None)
    return dict((node or {}).get("params") or {})


SAMPLE_DOCUMENTS = {
    "guideline-md": ("临床指南.md", "# 高血压指南\n高血压患者应规范随访，并依据风险分层选择治疗方案。"),
    "faq-csv": ("常见问题.csv", "问题,答案\n高血压是否需要随访,需要根据风险分层定期随访。"),
    "case-txt": ("病例摘要.txt", "患者出现高血压并伴随头痛，需要进一步评估。"),
}


def _preview_candidate(ref: str, params: dict[str, Any], value: dict[str, Any], index: int) -> dict[str, Any]:
    content = str(value.get("content") or value.get("text") or value.get("canonical_content") or "示例知识")
    source_chunk_id = str(value.get("source_chunk_id") or f"preview-chunk-{index + 1}")
    common = {
        "source_knowledge_id": f"preview-{ref}-{index + 1}", "source_chunk_id": source_chunk_id,
        "source_version_ids": [str(value.get("source_version_id") or "preview-version")],
        "source_anchor": f"{value.get('filename', '样例文档')}#chunk-{value.get('chunk_index', index)}",
        "anchor_json": dict(value.get("anchor") or {"chunk_index": value.get("chunk_index", index)}),
        "evidence_text": content, "is_primary": True,
    }
    kind = "qa" if ref in {"qa-generator", "multihop-qa"} else str(params.get("knowledge_type") or "text")
    mode = str(params.get("graph_mode") or "")
    if ref == "triple-builder" or kind == "graph" and mode != "semantic":
        data = {"subject": "高血压", "predicate": "需要", "object": "规范随访"}
        return {**common, "canonical_content": "高血压 需要 规范随访", "data_json": data}
    if ref == "evidence-binder" or kind == "graph" and mode == "semantic":
        data = {"source_entity": {"name": "高血压"}, "target_entity": {"name": "规范随访"},
                "relation": {"description": "患者需要"}, "evidence": [content]}
        return {**common, "canonical_content": "高血压 患者需要 规范随访", "data_json": data}
    if kind == "qa":
        data = {"question": "高血压患者需要什么？", "answer": "需要规范随访。"}
        return {**common, "canonical_content": f"{data['question']} {data['answer']}", "data_json": data}
    return {**common, "canonical_content": content, "data_json": {"content": content}}


def _preview_operator(ref: str, params: dict[str, Any], values: list[dict[str, Any]], root_documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Execute a deterministic, side-effect-free operator approximation for developer preview."""
    if ref in {"document-parser", "reviewed-source-chunk-input"}:
        return [dict(value) for value in root_documents]
    if ref in {"document-ir-normalizer", "null-filter", "language-filter", "text-cleaner", "whitespace-cleaner", "text-normalizer", "pii-compliance"}:
        result = []
        for raw in values:
            value = dict(raw); text = str(value.get("text", ""))
            if ref == "null-filter" and not text.strip():
                continue
            if ref in {"text-cleaner", "whitespace-cleaner", "text-normalizer"}:
                value["text"] = _clean_document_text(text, ref)
                if value.get("source_blocks"):
                    block_text, source_blocks = _clean_source_blocks(list(value["source_blocks"]), ref)
                    value["source_blocks"] = source_blocks
                    if value.get("source_type") == "docx":
                        value["text"] = block_text
                value["page_segments"] = [
                    {**segment, "text": _clean_document_text(str(segment.get("text", "")), ref)}
                    for segment in value.get("page_segments") or []
                ]
            result.append(value)
        return result
    if ref == "semantic-chunker":
        params = normalize_chunker_params(params)
        result = []
        for document in values:
            chunks = split_document_blocks(document, params) if document.get("source_blocks") else split_document_text(str(document.get("text", "")), params)
            for index, chunk in enumerate(chunks):
                anchor = dict(chunk.get("anchor") or {key: value for key, value in chunk.items() if key != "content"})
                result.append({"source_id": document.get("source_id", "preview-source"),
                               "source_version_id": document.get("source_version_id", "preview-version"),
                               "filename": document.get("filename", "样例文档"), "content": chunk["content"],
                               "chunk_index": index, "runtime_mode": "preview",
                               "anchor": {**dict(document.get("anchor") or {}), **anchor, "chunk_index": index}})
        return result
    if ref == "source-chunk-builder":
        return [{**value, "source_chunk_id": hashlib.sha256(f"preview:{value.get('source_version_id')}:{value.get('chunk_index')}".encode()).hexdigest()} for value in values]
    if ref == "text-knowledge-mapper":
        return [candidate for value in values if (candidate := _text_knowledge_candidate(value)) is not None]
    if ref == "entity-extractor":
        return [{**value, "entities": [
            {"name": "高血压", "type": "disease", "type_label": "疾病", "object_kind": "entity", "description": "常见心血管疾病", "aliases": [], "confidence": 0.9},
            {"name": "2500~3100 g", "object_kind": "literal", "literal_datatype": "range", "literal_unit": "g",
             "literal_raw_value": "2500~3100 g", "literal_normalized_value": {"min": 2500, "max": 3100}},
        ]} for value in values]
    if ref == "relation-extractor":
        return [{**value, "relations": [{"source": "高血压", "type": "uses_drug", "type_label": "使用药物", "target": "阿司匹林", "description": "", "keywords": [], "weight": None}]} for value in values]
    if ref == "literal-detector":
        return [_annotate_literals(value) for value in values]
    if ref == "entity-normalizer":
        return [{**value, "entities": [{**item, "entity_id": item.get("entity_id", f"preview-ent-{index}")}
                                       for index, item in enumerate((value.get("entities") or [])) if item.get("object_kind") != "literal"]} for value in values]
    if ref in {"triple-builder", "semantic-relation-builder", "evidence-binder"}:
        return [_preview_candidate(ref, params, value, index) for index, value in enumerate(values)]
    if ref in {"prompt-generator", "qa-generator", "graph-extractor", "structured-knowledge-generator", "multihop-qa"}:
        return [_preview_candidate(ref, params, value, index) for index, value in enumerate(values)]
    if ref == "graph-quality-validator":
        return [{**value, "graph_quality": {"hard_fail": False, "warnings": []}} for value in values]
    if ref == "artifact-merge":
        return [dict(value) for value in values]
    if ref == "quality-evaluator":
        rules = dict((params.get("_resolved_quality_profile") or {}).get("rules") or {})
        pass_score, review_score = float(rules.get("pass_score", 0.8)), float(rules.get("review_score", 0.6))
        return [{**value, "quality_score": float(value.get("quality_score", 1.0)),
                 "quality_status": "pass" if float(value.get("quality_score", 1.0)) >= pass_score
                 else "review" if float(value.get("quality_score", 1.0)) >= review_score else "reject"}
                for value in values]
    if ref == "quality-filter":
        rules = dict((params.get("_resolved_quality_profile") or {}).get("rules") or {})
        pass_score, review_score = float(rules.get("pass_score", 0.8)), float(rules.get("review_score", 0.6))
        return [{**value, "quality_status": "pass" if float(value.get("quality_score", 1.0)) >= pass_score else "review"}
                for value in values if float(value.get("quality_score", 1.0)) >= review_score]
    if ref == "deduplicate":
        unique: dict[str, dict[str, Any]] = {}
        for value in values:
            unique.setdefault(str(value.get("source_knowledge_id") or json.dumps(value, sort_keys=True, default=str)), dict(value))
        return list(unique.values())
    if ref in {"source-binding", "schema-validator", "knowledge-diff", "prompted-refiner"}:
        return [dict(value) for value in values]
    raise ValueError(f"算子不支持受控内存预览：{ref}")


def _truncate_preview_value(value: Any) -> tuple[Any, bool]:
    if isinstance(value, str):
        return (value[:500], len(value) > 500)
    if isinstance(value, list):
        items, flags = zip(*(_truncate_preview_value(item) for item in value)) if value else ((), ())
        return list(items), any(flags)
    if isinstance(value, dict):
        result: dict[str, Any] = {}; truncated = False
        for key, item in value.items():
            result[key], changed = _truncate_preview_value(item); truncated = truncated or changed
        return result, truncated
    return value, False


def _preview_port(values: list[dict[str, Any]]) -> dict[str, Any]:
    items = []; content_truncated = False
    for value in values[:3]:
        item, changed = _truncate_preview_value(value)
        items.append(item); content_truncated = content_truncated or changed
    return {"items": items, "total": len(values), "truncated": len(values) > 3 or content_truncated}


def _port_payload(values: list[dict[str, Any]], *, empty: bool = False) -> dict[str, Any]:
    return {} if empty else {"input": _preview_port(values)}


def _merge_preview_ports(ports: list[dict[str, Any]]) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    for payload in ports:
        for name, preview in payload.items():
            current = merged.setdefault(name, {"items": [], "total": 0, "truncated": False})
            current["total"] += int(preview.get("total", 0))
            current["truncated"] = current["truncated"] or bool(preview.get("truncated"))
            remaining = max(0, 3 - len(current["items"]))
            current["items"].extend(list(preview.get("items") or [])[:remaining])
            if current["total"] > len(current["items"]):
                current["truncated"] = True
    return merged


def preview_template_definition(definition: dict, sample_id: str, *, compiled_definition: dict | None = None) -> dict:
    """Preview a compiled DAG in memory without external calls or persistence."""
    if sample_id not in SAMPLE_DOCUMENTS:
        raise ValueError("不支持的内置样例")
    filename, text = SAMPLE_DOCUMENTS[sample_id]
    compiled = compiled_definition or definition
    root_documents = [{"source_id": "preview-source", "source_version_id": "preview-version", "filename": filename,
                       "text": text, "content": text, "source_chunk_id": "preview-reviewed-chunk", "chunk_index": 0,
                       "parser_strategy": "preview", "runtime_profile": "controlled_in_memory",
                       "parser_adapter": "preview", "anchor": {"file": filename, "page": None, "section": None}}]
    nodes = list(compiled.get("nodes") or [])
    incoming = _incoming(compiled)
    outputs: dict[str, list[dict[str, Any]]] = {}
    failures: dict[str, str] = {}
    expanded_runs: dict[str, dict[str, Any]] = {}
    for node in nodes:
        node_id = str(node["id"]); source_nodes = incoming.get(node_id, [])
        values = [value for source_id in source_nodes for value in outputs.get(source_id, [])]
        if node.get("kind") == "operator" and str(node.get("ref")) in {"document-parser", "reviewed-source-chunk-input"} and not source_nodes:
            values = [dict(value) for value in root_documents]
        failed_upstream = [source_id for source_id in source_nodes if source_id in failures]
        if failed_upstream:
            message = "上游节点失败，已跳过：" + "、".join(failed_upstream)
            failures[node_id] = message; outputs[node_id] = []
            expanded_runs[node_id] = {"status": "skipped", "inputs": _port_payload(values), "outputs": {}, "error": message}
            continue
        if node.get("kind") == "knowledge_sink":
            outputs[node_id] = []
            expanded_runs[node_id] = {"status": "success", "inputs": _port_payload(values), "outputs": {}, "error": None}
            continue
        try:
            result = _preview_operator(str(node.get("ref")), dict(node.get("params") or {}), values, root_documents)
            outputs[node_id] = result
            expanded_runs[node_id] = {"status": "success", "inputs": _port_payload(values), "outputs": {"output": _preview_port(result)}, "error": None}
        except Exception as exc:
            failures[node_id] = str(exc); outputs[node_id] = []
            expanded_runs[node_id] = {"status": "failed", "inputs": _port_payload(values), "outputs": {}, "error": str(exc)}

    top_level_runs: dict[str, dict[str, Any]] = {}
    compiled_edges = list(compiled.get("edges") or [])
    for node in definition.get("nodes") or []:
        node_id = str(node["id"])
        if node.get("kind") != "subflow":
            if node_id in expanded_runs:
                top_level_runs[node_id] = expanded_runs[node_id]
            continue
        prefix = f"{node_id}::"
        internal_ids = [str(item["id"]) for item in nodes if str(item["id"]).startswith(prefix)]
        internal_set = set(internal_ids)
        entries = [item for item in internal_ids if not any(edge["target"] == item and edge["source"] in internal_set for edge in compiled_edges)]
        exits = [item for item in internal_ids if not any(edge["source"] == item and edge["target"] in internal_set for edge in compiled_edges)]
        output_values = [value for item in exits for value in outputs.get(item, [])]
        internal_trace = {item[len(prefix):]: expanded_runs[item] for item in internal_ids if item in expanded_runs}
        statuses = {run["status"] for run in internal_trace.values()}
        status = "failed" if "failed" in statuses else "skipped" if "skipped" in statuses else "success"
        errors = [run["error"] for run in internal_trace.values() if run.get("error")]
        top_level_runs[node_id] = {"status": status, "inputs": _merge_preview_ports([expanded_runs.get(item, {}).get("inputs", {}) for item in entries]),
                                   "outputs": {"output": _preview_port(output_values)},
                                   "error": "；".join(errors) or None, "internal_trace": internal_trace}
    reviewed_root_ids = {str(node["id"]) for node in nodes if node.get("ref") == "reviewed-source-chunk-input"}
    chunk_values = [value for node_id, values in outputs.items()
                    if node_id.endswith("::chunk") or node_id in reviewed_root_ids for value in values]
    sink_totals = {str(node.get("output_key") or node.get("knowledge_type")): expanded_runs.get(str(node["id"]), {}).get("inputs", {}).get("input", {}).get("total", 0)
                   for node in nodes if node.get("kind") == "knowledge_sink"}
    return {"sample_id": sample_id, "filename": filename, "preview_mode": "controlled_in_memory",
            "status": "completed_with_errors" if failures else "completed", "node_runs": top_level_runs,
            "expanded_node_runs": expanded_runs, "stages": [str(node["id"]) for node in nodes],
            "parsed_text": text, "chunks": chunk_values, "outputs": sink_totals, "persisted": False}


def _source_key(source_id: str, output_type: str, anchor: str) -> str:
    return hashlib.sha256(f"{source_id}|{output_type}|{anchor}".encode("utf-8")).hexdigest()


def _initialize_llm_servings():
    """Validate the DataForge-owned Serving registry without reading secrets."""
    return get_llm_serving_registry()


def _llm_json(prompt: str, *, llm_serving: str = DEFAULT_LLM_SERVING_ID, system: str = "你是严谨的知识抽取器。只返回符合请求的 JSON 对象，不要输出 Markdown 或解释。") -> dict[str, Any]:
    """Call one configured Model Serving and parse its structured response."""
    registry = _initialize_llm_servings()
    serving, client = registry.client(llm_serving)
    extra_body: dict[str, Any] = {"app_id": "dataforge"}
    if serving.disable_thinking:
        extra_body["chat_template_kwargs"] = {"enable_thinking": False}
    started = time.monotonic()
    try:
        response = client.chat.completions.create(
            model=serving.model_name,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": prompt},
            ],
            response_format={"type": "json_object"},
            max_tokens=serving.max_tokens,
            extra_body=extra_body,
        )
    except APITimeoutError as exc:
        elapsed = time.monotonic() - started
        logger.error(
            "LLM Serving request timed out. serving_id=%s prompt_chars=%s elapsed_seconds=%.3f error_type=%s",
            serving.id, len(prompt), elapsed, type(exc).__name__,
        )
        raise TimeoutError(
            f"上游 LLM Serving {serving.id} 请求超时（{serving.timeout_seconds:g} 秒，未自动重试）"
        ) from exc
    except Exception as exc:
        elapsed = time.monotonic() - started
        logger.error(
            "LLM Serving request failed. serving_id=%s prompt_chars=%s elapsed_seconds=%.3f error_type=%s",
            serving.id, len(prompt), elapsed, type(exc).__name__,
        )
        raise RuntimeError(f"上游 LLM Serving {serving.id} 调用失败（{type(exc).__name__}）") from exc
    try:
        content = response.choices[0].message.content if response.choices else ""
        if not isinstance(content, str) or not content.strip():
            raise ValueError("LLM 没有返回 JSON 内容")
        parsed = json.loads(content)
    except (AttributeError, IndexError, TypeError, ValueError) as exc:
        logger.error(
            "LLM Serving response invalid. serving_id=%s prompt_chars=%s elapsed_seconds=%.3f error_type=%s",
            serving.id, len(prompt), time.monotonic() - started, type(exc).__name__,
        )
        if isinstance(exc, json.JSONDecodeError):
            raise ValueError(f"LLM 返回的内容不是 JSON：{exc.msg}") from exc
        raise
    logger.info(
        "LLM Serving request completed. serving_id=%s prompt_chars=%s elapsed_seconds=%.3f",
        serving.id, len(prompt), time.monotonic() - started,
    )
    return parsed


def _json_errors(data: dict[str, Any], schema: dict[str, Any]) -> list[str]:
    try:
        from jsonschema import Draft202012Validator
        return [error.message for error in Draft202012Validator(schema).iter_errors(data)]
    except ImportError:
        return [f"缺少必填字段 {item}" for item in schema.get("required", []) if not data.get(item)]


def _chunk_prompt(output_type: str, contract: dict[str, Any], content: str) -> str:
    schema = contract["schema"]
    instructions = {
        "qa": "从当前来源分块生成全部有事实依据且可独立回答的问答对。每项必须包含 question 和 answer。",
        "graph": "从当前来源分块抽取全部明确陈述的关系三元组。每项必须包含 subject、predicate、object。",
        "graph:triple": "从当前来源分块抽取明确关系。每项包含 subject、predicate、object。",
        "graph:semantic": "抽取语义实体关系。每项包含 source_entity、target_entity、relation 和可核验 evidence。",
    }
    type_instruction = instructions.get(output_type, "")
    template = contract.get("prompt") or "根据当前来源分块生成全部有效的结构化知识项。"
    return (
        f"{type_instruction}\n{template}\n\n"
        "仅使用下方当前来源分块；没有有效结果时返回 {\"items\": []}。"
        "必须返回 JSON 对象，且 items 是数组。items 中每一项都必须符合此 JSON Schema：\n"
        f"{json.dumps(schema, ensure_ascii=False)}\n\n"
        f"当前来源分块：\n{content}"
    )


def _item_errors(items: Any, schema: dict[str, Any]) -> list[str]:
    if not isinstance(items, list):
        return ["顶层 JSON 必须包含 items 数组"]
    errors: list[str] = []
    for index, data in enumerate(items):
        if not isinstance(data, dict):
            errors.append(f"items[{index}]：必须是 JSON 对象")
            continue
        errors.extend(f"items[{index}]：{message}" for message in _json_errors(data, schema))
    return errors


def _structured_candidates(source: Source, version: SourceVersion, output_type: str, chunk: dict[str, Any], contract: dict[str, Any], *, llm_serving: str = DEFAULT_LLM_SERVING_ID) -> list[dict]:
    prompt = _chunk_prompt(output_type, contract, str(chunk["content"]))
    response = _llm_json(prompt, llm_serving=llm_serving)
    items = response.get("items") if isinstance(response, dict) else None
    errors = _item_errors(items, contract["schema"])
    if errors:
        repair = prompt + "\n\n上次输出校验失败，请只返回修复后的 JSON。错误：" + "；".join(errors)
        response = _llm_json(repair, llm_serving=llm_serving)
        items = response.get("items") if isinstance(response, dict) else None
        errors = _item_errors(items, contract["schema"])
    if errors:
        raise ValueError("LLM 一次修复后仍未通过 Schema 校验：" + "；".join(errors))
    anchor = {"file": source.original_filename, "chunk_index": int(chunk["chunk_index"]), "source_chunk_id": str(chunk["source_chunk_id"])}
    result: list[dict] = []
    for data in items:
        if output_type == "graph:semantic":
            data = dict(data)
            data["evidence"] = [{"source_version_id": version.id, "source_chunk_id": str(chunk["source_chunk_id"])}]
        def nested(path: str) -> Any:
            current: Any = data
            for part in path.split("."):
                current = current.get(part) if isinstance(current, dict) else None
            return current
        canonical = " ".join(str(nested(path) or "").strip() for path in contract.get("canonical_fields") or [contract["canonical_field"]]).strip()
        if not canonical:
            raise ValueError(f"LLM 输出缺少 canonical 字段 {contract['canonical_field']}")
        identity = "|".join(str(nested(field) or "") for field in contract.get("identity_fields", [])) or canonical
        result.append({
            "source_knowledge_id": _source_key(source.id, output_type, f"{chunk['chunk_index']}|{identity}"),
            "canonical_content": canonical,
            "data_json": data,
            "source_version_ids": [version.id],
            "source_chunk_id": chunk["source_chunk_id"],
            "source_anchor": f"{source.original_filename}#chunk-{chunk['chunk_index']}",
            "anchor_json": anchor,
            "evidence_text": chunk["content"],
            "is_primary": True,
        })
    return result


def _text_knowledge_candidate(chunk: dict[str, Any]) -> dict[str, Any] | None:
    """Map one frozen reviewed chunk without normalizing its body or provenance."""
    content = chunk.get("content")
    if content is None:
        return None
    if not isinstance(content, str):
        raise ValueError("文本知识映射器要求 content 为字符串")
    if not content.strip():
        return None
    data = {"filename": chunk["filename"], "chunk_index": chunk["chunk_index"]}
    for key in ("heading_context", "metadata"):
        if key in chunk:
            data[key] = deepcopy(chunk[key])
    anchor = deepcopy(chunk.get("anchor") or {})
    anchor.update(file=chunk["filename"], chunk_index=chunk["chunk_index"], source_chunk_id=chunk["source_chunk_id"])
    return {
        "source_knowledge_id": _source_key(chunk["source_id"], "text", str(chunk["chunk_index"])),
        "canonical_content": content, "data_json": data,
        "source_version_ids": [chunk["source_version_id"]], "source_chunk_id": chunk["source_chunk_id"],
        "source_anchor": f"{chunk['filename']}#chunk-{chunk['chunk_index']}",
        "anchor_json": anchor, "evidence_text": content, "is_primary": True,
        **{key: chunk[key] for key in ("source_chunk_revision_id", "source_review_snapshot_id") if key in chunk},
    }


def _generated_text_candidates(source: Source, version: SourceVersion, chunk: dict[str, Any],
                               contract: dict[str, Any], params: dict[str, Any], llm_serving: str) -> list[dict[str, Any]]:
    """v6 text generation uses the node's frozen prompt, never the type-wide fallback."""
    frozen_prompt = params.get("_resolved_prompt_template") or {}
    if not str(frozen_prompt.get("body") or "").strip():
        raise ValueError("文本生成缺少节点冻结的 Prompt Template Revision")
    mapped = _text_knowledge_candidate(chunk)
    if mapped is None:
        return []
    effective_contract = {
        **contract, "prompt": frozen_prompt["body"],
        "schema": {"allOf": [contract["schema"], {
            "type": "object", "required": ["canonical_content"],
            "properties": {"canonical_content": {"type": "string", "pattern": r"\S"}},
        }]},
        "canonical_field": "canonical_content", "canonical_fields": ["canonical_content"],
        # Candidate identity and provenance are server-owned, not model fields.
        "identity_fields": [],
    }
    candidates = _structured_candidates(source, version, "text", chunk, effective_contract, llm_serving=llm_serving)
    reserved = {"source_knowledge_id", "source_id", "source_version_id", "source_version_ids", "source_chunk_id",
                "source_chunk_revision_id", "source_review_snapshot_id", "source_anchor", "anchor", "anchor_json",
                "evidence_text", "evidence", "is_primary"}
    return [{**deepcopy(mapped), "source_knowledge_id": candidate["source_knowledge_id"],
             "canonical_content": candidate["canonical_content"],
             "data_json": {**{key: value for key, value in candidate["data_json"].items() if key not in reserved},
                           **deepcopy(mapped["data_json"])}}
            for candidate in candidates]


def _candidates(source: Source, version: SourceVersion, output_type: str, chunk: dict[str, Any], *, contract: dict[str, Any] | None = None, llm_serving: str = DEFAULT_LLM_SERVING_ID) -> list[dict]:
    anchor = {"file": source.original_filename, "chunk_index": int(chunk["chunk_index"]), "source_chunk_id": str(chunk["source_chunk_id"])}
    if output_type == "text":
        return [{"source_knowledge_id": _source_key(source.id, "text", str(chunk["chunk_index"])), "canonical_content": chunk["content"], "data_json": {"filename": source.original_filename, "chunk_index": chunk["chunk_index"]}, "source_version_ids": [version.id], "source_chunk_id": chunk["source_chunk_id"], "source_anchor": f"{source.original_filename}#chunk-{chunk['chunk_index']}", "anchor_json": anchor, "evidence_text": chunk["content"], "is_primary": True}]
    if not contract:
        raise ValueError(f"不支持的知识类型或缺少已发布契约：{output_type}")
    return _structured_candidates(source, version, output_type, chunk, contract, llm_serving=llm_serving)


def _incoming(definition: dict[str, Any]) -> dict[str, list[str]]:
    values: dict[str, list[str]] = {node["id"]: [] for node in definition.get("nodes", [])}
    for edge in definition.get("edges", []):
        values.setdefault(edge["target"], []).append(edge["source"])
    return values


def select_parser_adapter(filename: str, profile: str, environ: dict[str, str] | None = None) -> str:
    """Resolve deployment-only MinerU variants behind one logical parser node."""
    variables = environ if environ is not None else os.environ
    suffix = Path(filename).suffix.lower()
    if suffix in {".doc", ".docx"}:
        return "dataforge-word-parser"
    if suffix in {".csv", ".xlsx"}:
        return "dataforge-structured-table-parser"
    if suffix in {".md", ".txt"}:
        return "dataforge-text-parser"
    if suffix == ".pdf":
        return "mineru-pipeline-gpu"
    raise ValueError("Document Parser 不支持的文档类型")


def select_runtime_mode(record_count: int, environ: dict[str, str] | None = None) -> str:
    """Choose normal vs. DataFlow Batch adapter without changing the Flow DSL."""
    variables = environ if environ is not None else os.environ
    try:
        threshold = max(2, int(variables.get("DATAFORGE_BATCH_THRESHOLD", "32")))
    except ValueError as exc:
        raise ValueError("DATAFORGE_BATCH_THRESHOLD 必须是整数") from exc
    return "batch" if record_count >= threshold else "single"


def _documents_for_versions(objects, versions: list[SourceVersion], sources: dict[str, Source], flow_run_id: str,
                            created_parser_keys: list[str], *, force_ocr: bool = False) -> list[dict[str, Any]]:
    documents: list[dict[str, Any]] = []
    for version in versions:
        source = sources[version.source_id]
        payload = objects.get_bytes(version.object_key)
        suffix = Path(source.original_filename).suffix.lower()
        parser_artifacts: list[dict[str, Any]] = []
        page_segments: list[dict[str, Any]] = []
        source_blocks: list[dict[str, Any]] = []
        source_type: str | None = None
        if suffix == ".pdf":
            parsed = parse_with_mineru(filename=source.original_filename, payload=payload, parse_method="ocr" if force_ocr else "auto")
            text = parsed.markdown
            source_blocks = content_list_blocks(parsed.content_list)
            page_segments = _page_segments_from_blocks(source_blocks)
            source_type = "pdf"
            encoded_middle = json.dumps(parsed.middle_json, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
            object_key = f"parser-artifacts/{version.id}/{flow_run_id}/mineru-middle.json"
            stored = objects.put_bytes(object_key, encoded_middle, "application/json")
            created_parser_keys.append(stored.key)
            pdf_info = parsed.middle_json.get("pdf_info")
            page_count = len(pdf_info) if isinstance(pdf_info, list) else len(page_segments)
            parser_artifacts.append({
                "type_code": "parser.middle-json", "uri": f"object:///{stored.key}", "checksum": stored.sha256,
                "source_version_id": version.id,
                "data": {"object_key": stored.key, "mineru_version": parsed.version, "backend": parsed.backend,
                         "parse_method": parsed.parse_method, "lang_list": "ch", "formula_enable": True,
                         "table_enable": True, "page_count": page_count, "size_bytes": stored.size_bytes},
            })
            profile = "pipeline:auto"
        else:
            if suffix == ".docx":
                text, source_blocks = _docx_source_blocks(payload)
                source_type = "docx"
            else:
                text = _extract_text(source.original_filename, payload)
            profile = "native"
        faq_table_rows = parse_table_rows(source.original_filename, payload) \
            if FAQ_FILENAME_PATTERN.fullmatch(Path(source.original_filename).name) else []
        documents.append({"source_id": source.id, "source_version_id": version.id, "filename": source.original_filename,
                          "text": text, "parser_strategy": "auto", "runtime_profile": profile,
                          "parser_adapter": select_parser_adapter(source.original_filename, profile),
                          "page_segments": page_segments, "source_blocks": source_blocks,
                          "source_type": source_type, "_parser_artifacts": parser_artifacts,
                          "_faq_table_rows": faq_table_rows,
                          "anchor": {"file": source.original_filename, "page": None, "section": None,
                                     "anchor_version": 2 if source_type else 1,
                                     "source_version_id": version.id, "source_type": source_type,
                                     "blocks": source_blocks if source_type else []}})
    return documents


def _normalized_name(value: str) -> str:
    """Normalize an entity name for identity and merging."""
    return re.sub(r"[\s　]+", "", str(value or "")).casefold().strip("。，,;；:：()（）[]《》")


def _graph_output_key(params: dict[str, Any]) -> str:
    mode = str(params.get("graph_mode") or "")
    return f"graph:{mode}" if mode else "graph"


def _graph_config_from_contracts(type_contracts: dict[str, dict[str, Any]]) -> GraphExtractionConfig:
    for contract in type_contracts.values():
        if contract.get("graph_config") is not None:
            return normalize_graph_config(contract["graph_config"])
    return normalize_graph_config(None)


def _resolve_type(config: GraphExtractionConfig, raw_type: Any, *, relation: bool = False) -> tuple[str, str] | None:
    """Resolve an LLM-reported type to ``(code, label)`` when it matches the schema."""
    if not raw_type:
        return None
    raw = str(raw_type).strip()
    definitions = config.relation_types if relation else config.entity_types
    for item in definitions:
        if item.code == raw:
            return item.code, item.label
    for item in definitions:
        if item.label == raw:
            return item.code, item.label
    return None


def _graph_entity_id(library_id: str, type_code: str | None, name: str) -> str:
    return "ent_" + hashlib.sha256(f"{library_id}|{type_code or ''}|{_normalized_name(name)}".encode("utf-8")).hexdigest()[:24]


def _graph_relation_id(library_id: str, source_id: str, rel_type: str, target_id: str) -> str:
    return "rel_" + hashlib.sha256(f"{library_id}|{source_id}|{rel_type}|{target_id}".encode("utf-8")).hexdigest()[:24]


def _graph_knowledge_id(library_id: str, graph_mode: str, *parts: str) -> str:
    return hashlib.sha256("|".join((library_id, graph_mode, *parts)).encode("utf-8")).hexdigest()


def _literal_entity(name: str) -> dict[str, Any] | None:
    literal = detect_literal(name)
    if literal is None:
        return None
    return {
        "name": name, "type": None, "type_label": None, "object_kind": "literal",
        "literal_datatype": literal.datatype, "literal_unit": literal.unit,
        "literal_raw_value": literal.raw_value, "literal_normalized_value": literal.normalized_value,
        "description": "", "aliases": [], "confidence": 1.0,
    }


def _graph_config_for_node(config: GraphExtractionConfig, params: dict[str, Any], *, relation: bool = False) -> GraphExtractionConfig:
    raw = config.to_dict()
    selected_entities = {str(item) for item in (params.get("entity_types") or []) if str(item)}
    if selected_entities and params.get("entity_type_scope") != "all":
        raw["entity_types"] = [item for item in raw["entity_types"] if item["code"] in selected_entities]
        # Entity extraction does not consume relation definitions. Keeping them
        # here would leave references to types excluded by this node's subset.
        if not relation:
            raw["relation_types"] = []
    selected_relations = {str(item) for item in (params.get("relation_types") or []) if str(item)}
    if selected_relations:
        raw["relation_types"] = [item for item in raw["relation_types"] if item["code"] in selected_relations]
    if relation and params.get("unknown_relation_policy"):
        raw["unknown_relation_policy"] = params["unknown_relation_policy"]
    if not relation and params.get("unknown_entity_policy"):
        raw["unknown_entity_policy"] = params["unknown_entity_policy"]
    if not relation and params.get("prompt_mode"):
        raw.setdefault("prompt", {})["mode"] = params["prompt_mode"]
    constraints = {str(item.get("relation_type")): item for item in (params.get("relation_constraints") or [])
                   if isinstance(item, dict) and item.get("relation_type")}
    for item in raw.get("relation_types") or []:
        constraint = constraints.get(str(item.get("code")))
        if constraint:
            item["source_types"] = list(constraint.get("source_types") or [])
            item["target_types"] = list(constraint.get("target_types") or [])
    return normalize_graph_config(raw)


def _extract_entities(chunk: dict[str, Any], config: GraphExtractionConfig, llm_serving: str,
                      params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    """One LLM call extracting typed entities from a single source chunk."""
    if (params or {}).get("entity_type_scope") == "subset" and not params.get("entity_types"):
        return []
    system, prompt = entity_prompt_for(config, str(chunk.get("content", "")))
    response = _llm_json(prompt, llm_serving=llm_serving, system=system)
    raw_entities = response.get("entities") if isinstance(response, dict) else None
    if not isinstance(raw_entities, list):
        raise ValueError("LLM 实体抽取未返回 entities 数组")
    params = dict(params or {})
    entities: list[dict[str, Any]] = []
    for raw in raw_entities:
        if not isinstance(raw, dict):
            continue
        name = str(raw.get("name") or "").strip()
        if not name:
            continue
        literal_entity = _literal_entity(name)
        if literal_entity is not None:
            entities.append(literal_entity)
            continue
        resolved = _resolve_type(config, raw.get("type"))
        if resolved is None:
            # An empty entity_types schema is "unconstrained": keep the LLM's
            # free-form type code instead of rejecting every entity.
            if config.entity_types and config.unknown_entity_policy == "reject":
                continue  # schema-defined templates reject unknown types, never "未分类"
            code = str(raw.get("type") or "").strip() or "other"
            label = code
        else:
            code, label = resolved
        confidence = float(raw.get("confidence") or 1.0)
        if confidence < float(params.get("confidence_threshold", 0.0)):
            continue
        entities.append({
            "name": name, "type": code, "type_label": label,
            "description": str(raw.get("description") or "").strip() if params.get("generate_description", True) else "",
            "aliases": ([str(item).strip() for item in (raw.get("aliases") or []) if str(item).strip()]
                        if params.get("extract_aliases", True) else []),
            "object_kind": "entity", "confidence": confidence,
        })
    return entities


def _extract_relations(chunk: dict[str, Any], entities: list[dict[str, Any]], config: GraphExtractionConfig, llm_serving: str) -> list[dict[str, Any]]:
    """One LLM call extracting relations between already-extracted entities."""
    entity_names = [item["name"] for item in entities if item.get("object_kind") != "literal"]
    system, prompt = relation_prompt_for(config, entity_names, str(chunk.get("content", "")))
    response = _llm_json(prompt, llm_serving=llm_serving, system=system)
    raw_relations = response.get("relations") if isinstance(response, dict) else None
    if not isinstance(raw_relations, list):
        raise ValueError("LLM 关系抽取未返回 relations 数组")
    relations: list[dict[str, Any]] = []
    for raw in raw_relations:
        if not isinstance(raw, dict):
            continue
        source = str(raw.get("source") or "").strip()
        target = str(raw.get("target") or "").strip()
        if not source or not target:
            continue
        resolved = _resolve_type(config, raw.get("type"), relation=True)
        if resolved is None:
            # An empty relation_types schema is "unconstrained": keep the LLM's
            # free-form source-language relation type instead of rejecting every relation.
            if config.relation_types and config.unknown_relation_policy == "reject":
                continue
            code = str(raw.get("type") or "").strip() or "other"
            label = str(raw.get("label") or code).strip()
        else:
            code, schema_label = resolved
            # The schema code remains the stable identity; the extracted label
            # is the source-language relationship shown to business users.
            label = str(raw.get("label") or schema_label).strip()
        definition = config.relation_by_code(code)
        if definition and (definition.source_types or definition.target_types):
            entity_types = {str(item.get("name")): str(item.get("type") or "") for item in entities}
            if definition.source_types and entity_types.get(source) not in definition.source_types:
                continue
            if definition.target_types and entity_types.get(target) not in definition.target_types:
                continue
        relations.append({
            "source": source, "target": target, "type": code, "type_label": label,
            "description": str(raw.get("description") or "").strip(),
            "keywords": [str(item).strip() for item in (raw.get("keywords") or []) if str(item).strip()],
            "weight": float(raw["weight"]) if raw.get("weight") not in (None, "") else None,
        })
    return relations


def _annotate_literals(record: dict[str, Any]) -> dict[str, Any]:
    """Literal Detector: rule classification is already applied in extraction; re-assert here."""
    value = dict(record)
    entities = []
    for item in value.get("entities") or []:
        item = dict(item)
        if item.get("object_kind") != "literal":
            literal_entity = _literal_entity(str(item.get("name") or ""))
            if literal_entity is not None:
                entities.append(literal_entity)
                continue
        entities.append(item)
    value["entities"] = entities
    return value


def _normalize_entities(record: dict[str, Any], library_id: str, config: GraphExtractionConfig) -> dict[str, Any]:
    """Entity Normalizer: drop literals (semantic never uses them as nodes) and merge aliases."""
    value = dict(record)
    merged: list[dict[str, Any]] = []
    by_name: dict[str, dict[str, Any]] = {}
    for item in value.get("entities") or []:
        item = dict(item)
        if item.get("object_kind") == "literal":
            continue  # semantic literals leave the entity set; they are facts, not nodes
        name = str(item.get("name") or "").strip()
        if not name:
            continue
        key = _normalized_name(name)
        existing = by_name.get(key)
        if existing is None:
            by_name[key] = item
            merged.append(item)
            continue
        existing["aliases"] = sorted(set((existing.get("aliases") or []) + [name] + (item.get("aliases") or [])))
        existing["confidence"] = max(float(existing.get("confidence") or 0), float(item.get("confidence") or 0))
    for item in merged:
        item["entity_id"] = _graph_entity_id(library_id, item.get("type"), item["name"])
    value["entities"] = merged
    return value


def _candidate_meta(source: Source, version: SourceVersion, chunk: dict[str, Any], canonical: str, data_json: dict[str, Any]) -> dict[str, Any]:
    anchor = {"file": source.original_filename, "chunk_index": int(chunk["chunk_index"]), "source_chunk_id": str(chunk["source_chunk_id"])}
    return {
        "canonical_content": canonical,
        "data_json": data_json,
        "source_version_ids": [version.id],
        "source_chunk_id": chunk["source_chunk_id"],
        "source_anchor": f"{source.original_filename}#chunk-{chunk['chunk_index']}",
        "anchor_json": anchor,
        "evidence_text": chunk["content"],
        "is_primary": True,
    }


def _build_triples(record: dict[str, Any], config: GraphExtractionConfig, library_id: str, sources: dict[str, Source], versions: dict[str, SourceVersion]) -> list[dict[str, Any]]:
    """Triple Builder: assemble entity→entity relations and entity→literal facts."""
    entities = {_normalized_name(item["name"]): item for item in record.get("entities") or []}
    source = sources[record["source_id"]]
    version = versions[record["source_version_id"]]
    candidates: list[dict[str, Any]] = []
    for rel in record.get("relations") or []:
        subject_name = str(rel.get("source") or "").strip()
        target_text = str(rel.get("target") or "").strip()
        if not subject_name or not target_text:
            continue
        subject = entities.get(_normalized_name(subject_name)) or {"name": subject_name, "type": None, "type_label": None, "object_kind": "entity"}
        target = entities.get(_normalized_name(target_text))
        literal = detect_literal(target_text) if (target is None or target.get("object_kind") != "entity") else None
        predicate_code = str(rel.get("type") or "")
        predicate_label = str(rel.get("type_label") or rel.get("type") or "")
        if target is not None and target.get("object_kind") == "literal":
            data = {"object_kind": "literal", "literal_datatype": target.get("literal_datatype"),
                    "literal_unit": target.get("literal_unit"), "literal_raw_value": target.get("literal_raw_value", target_text),
                    "literal_normalized_value": target.get("literal_normalized_value")}
            object_type = None
            identity = ("literal", _normalized_name(subject_name), predicate_code, str(target.get("literal_normalized_value") or target_text), str(target.get("literal_unit") or ""))
        elif literal is not None:
            data = {"object_kind": "literal", "literal_datatype": literal.datatype, "literal_unit": literal.unit,
                    "literal_raw_value": literal.raw_value, "literal_normalized_value": literal.normalized_value}
            object_type = None
            identity = ("literal", _normalized_name(subject_name), predicate_code, str(literal.normalized_value), str(literal.unit or ""))
        else:
            data = {"object_kind": "entity"}
            object_type = (target or {}).get("type")
            identity = ("entity", _normalized_name(subject_name), predicate_code, _normalized_name(target_text))
        data_json: dict[str, Any] = {
            "subject": subject_name, "predicate": predicate_label, "predicate_code": predicate_code or None,
            "object": target_text, "data": data,
        }
        if subject.get("type"):
            data_json["subject_type"] = subject.get("type")
            data_json["subject_type_label"] = subject.get("type_label")
        if object_type:
            data_json["object_type"] = object_type
        canonical = " ".join(str(part).strip() for part in (subject_name, predicate_label or predicate_code, target_text) if part)
        candidate = _candidate_meta(source, version, record, canonical, data_json)
        candidate["source_knowledge_id"] = _graph_knowledge_id(library_id, "triple", *identity)
        candidates.append(candidate)
    return candidates


def _build_semantic_relations(record: dict[str, Any], config: GraphExtractionConfig, library_id: str, sources: dict[str, Source], versions: dict[str, SourceVersion]) -> list[dict[str, Any]]:
    """Semantic Relation Builder: only Entity → Entity, with type + description."""
    entities = {_normalized_name(item["name"]): item for item in record.get("entities") or []}
    source = sources[record["source_id"]]
    version = versions[record["source_version_id"]]
    candidates: list[dict[str, Any]] = []
    for rel in record.get("relations") or []:
        source_name = str(rel.get("source") or "").strip()
        target_name = str(rel.get("target") or "").strip()
        source_entity = entities.get(_normalized_name(source_name))
        target_entity = entities.get(_normalized_name(target_name))
        if source_entity is None or target_entity is None:
            continue  # semantic endpoints must be known entities, never literals
        if detect_literal(source_name) is not None or detect_literal(target_name) is not None:
            continue
        relation_type = str(rel.get("type") or "")
        source_id = source_entity.get("entity_id") or _graph_entity_id(library_id, source_entity.get("type"), source_name)
        target_id = target_entity.get("entity_id") or _graph_entity_id(library_id, target_entity.get("type"), target_name)
        relation_id = _graph_relation_id(library_id, source_id, relation_type, target_id)
        data_json = {
            "source_entity": {"entity_id": source_id, "name": source_name, "type": source_entity.get("type"),
                              "type_label": source_entity.get("type_label"), "description": source_entity.get("description") or "",
                              "aliases": source_entity.get("aliases") or [], "confidence": source_entity.get("confidence")},
            "target_entity": {"entity_id": target_id, "name": target_name, "type": target_entity.get("type"),
                              "type_label": target_entity.get("type_label"), "description": target_entity.get("description") or "",
                              "aliases": target_entity.get("aliases") or [], "confidence": target_entity.get("confidence")},
            "relation": {"relation_id": relation_id, "type": relation_type, "type_label": rel.get("type_label"),
                         "description": rel.get("description") or "", "keywords": rel.get("keywords") or [], "weight": rel.get("weight")},
        }
        canonical = " ".join(str(part).strip() for part in (source_name, rel.get("type_label") or relation_type, target_name) if part)
        candidate = _candidate_meta(source, version, record, canonical, data_json)
        candidate["source_knowledge_id"] = _graph_knowledge_id(library_id, "semantic", source_id, relation_type, target_id)
        candidates.append(candidate)
    return candidates


def _bind_evidence(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Evidence Binder: attach the source chunk reference to every semantic relation."""
    result = []
    for candidate in candidates:
        candidate = dict(candidate)
        data = dict(candidate.get("data_json") or {})
        data["evidence"] = [{"source_version_id": (candidate.get("source_version_ids") or [None])[0],
                             "source_chunk_id": str(candidate.get("source_chunk_id") or "")}]
        candidate["data_json"] = data
        result.append(candidate)
    return result


def _validate_graph_item(data: dict[str, Any], config: GraphExtractionConfig, graph_mode: str) -> None:
    """Schema Validator: enforce the template graph schema on one assembled item."""
    entity_codes = config.entity_codes()
    relation_codes = {item.code for item in config.relation_types}
    if graph_mode == "semantic":
        source, target, relation = (data.get(key) or {} for key in ("source_entity", "target_entity", "relation"))
        for role, entity in (("source_entity", source), ("target_entity", target)):
            if not entity.get("name"):
                raise ValueError("语义实体缺少 name")
            if not entity.get("description"):
                raise ValueError(f"{role} 缺少 description")
            if entity_codes and entity.get("type") not in entity_codes:
                raise ValueError(f"{role} 实体类型非法：{entity.get('type')}")
            if detect_literal(str(entity.get("name"))) is not None:
                raise ValueError(f"{role} 是字面值，禁止作为语义实体节点")
        if relation_codes and relation.get("type") not in relation_codes:
            raise ValueError(f"关系类型非法：{relation.get('type')}")
        relation_def = config.relation_by_code(str(relation.get("type") or ""))
        if relation_def is not None:
            if relation_def.source_types and source.get("type") not in relation_def.source_types:
                raise ValueError(f"关系 {relation_def.code} 不允许 source 类型 {source.get('type')}")
            if relation_def.target_types and target.get("type") not in relation_def.target_types:
                raise ValueError(f"关系 {relation_def.code} 不允许 target 类型 {target.get('type')}")
        if not relation.get("description"):
            raise ValueError("语义关系缺少 description")
        if not data.get("evidence"):
            raise ValueError("语义关系缺少 Evidence")
    else:
        if entity_codes and data.get("subject_type") not in entity_codes:
            raise ValueError(f"三元组 subject_type 非法：{data.get('subject_type')}")
        object_kind = (data.get("data") or {}).get("object_kind")
        if object_kind != "literal":
            if entity_codes and data.get("object_type") not in entity_codes:
                raise ValueError(f"三元组 object_type 非法：{data.get('object_type')}")


def _run_operator(ref: str, params: dict[str, Any], values: list[dict[str, Any]], *, root_documents: list[dict[str, Any]], sources: dict[str, Source], versions: dict[str, SourceVersion], type_contracts: dict[str, dict[str, Any]], job_id: str | None = None, store: V7Store | None = None, retry_scope: set[tuple[str, str, str]] | None = None, generation: dict[str, dict[str, list[dict[str, Any]]]] | None = None, graph_config: GraphExtractionConfig | None = None, sink_libraries: dict[str, str] | None = None, operator_version: int | None = None) -> list[dict[str, Any]]:
    """Execute only DataForge adapters; DataFlow class names never enter a Flow."""
    cfg = graph_config or _graph_config_from_contracts(type_contracts)
    library_id = str((sink_libraries or {}).get(_graph_output_key(params)) or "")
    if ref == "document-parser":
        return root_documents
    if ref == "reviewed-source-chunk-input":
        return [dict(value) for value in root_documents]
    if ref == "text-knowledge-mapper":
        outcome = (generation if generation is not None else {}).setdefault("text", {"successful": [], "failed": [], "targeted": []})
        result = []
        for chunk in values:
            scope_key = ("text", str(chunk["source_version_id"]), str(chunk["source_chunk_id"]))
            if retry_scope is not None and scope_key not in retry_scope:
                continue
            outcome["targeted"].append(chunk)
            try:
                candidate = _text_knowledge_candidate(chunk)
                if candidate is not None:
                    result.append(candidate)
                outcome["successful"].append(chunk)
                if store and job_id:
                    store.record_chunk_generation(job_id, "text", chunk, status="completed", candidate_count=int(candidate is not None))
            except Exception as exc:
                outcome["failed"].append({**chunk, "error": str(exc)})
                if store and job_id:
                    store.record_chunk_generation(job_id, "text", chunk, status="failed", error=str(exc))
        return result
    if ref in {"document-ir-normalizer", "null-filter", "language-filter", "text-cleaner", "whitespace-cleaner", "text-normalizer", "pii-compliance"}:
        result = []
        mode = select_runtime_mode(len(values))
        for value in values:
            text = str(value.get("text", ""))
            if ref == "null-filter" and not text.strip():
                continue
            value = dict(value)
            if ref in {"text-cleaner", "whitespace-cleaner", "text-normalizer"}:
                value["text"] = _clean_document_text(text, ref)
                if value.get("source_blocks"):
                    block_text, source_blocks = _clean_source_blocks(list(value["source_blocks"]), ref)
                    value["source_blocks"] = source_blocks
                    if value.get("source_type") == "docx":
                        value["text"] = block_text
                    if value.get("source_type") == "pdf":
                        value["page_segments"] = _page_segments_from_blocks(source_blocks)
                value["page_segments"] = [
                    {**segment, "text": _clean_document_text(str(segment.get("text", "")), ref)}
                    for segment in value.get("page_segments") or []
                ]
            if ref == "text-cleaner":
                value["runtime_mode"] = mode
            result.append(value)
        return result
    if ref == "semantic-chunker":
        params = normalize_chunker_params(params)
        result = []; mode = select_runtime_mode(len(values))
        for document in values:
            if document.get("source_blocks"):
                for index, chunk in enumerate(split_document_blocks(document, params)):
                    result.append({"source_id": document["source_id"], "source_version_id": document["source_version_id"],
                                   "filename": document["filename"], "content": chunk["content"], "chunk_index": index,
                                   "runtime_mode": mode, "anchor": {**dict(chunk["anchor"]), "chunk_index": index}})
                continue
            segments = document.get("page_segments") or [{"text": str(document.get("text", "")), "page": None, "page_index": None}]
            if not params["preserve_page_boundary"]:
                segments = [{
                    "text": "\n\n".join(str(item.get("text", "")) for item in segments),
                    "page": segments[0].get("page") if len(segments) == 1 else None,
                    "page_index": segments[0].get("page_index") if len(segments) == 1 else None,
                    "page_start": segments[0].get("page") if segments else None,
                    "page_end": segments[-1].get("page") if segments else None,
                }]
            index = 0
            for segment in segments:
                for chunk in split_document_text(str(segment.get("text", "")), params):
                    result.append({"source_id": document["source_id"], "source_version_id": document["source_version_id"],
                                   "filename": document["filename"], "content": chunk["content"], "chunk_index": index,
                                   "runtime_mode": mode, "anchor": {**document.get("anchor", {}), "page": segment.get("page"),
                                                                       "page_index": segment.get("page_index"),
                                                                       "page_start": segment.get("page_start"), "page_end": segment.get("page_end"),
                                                                       **{key: value for key, value in chunk.items() if key != "content"},
                                                                       "chunk_index": index}})
                    index += 1
        return result
    if ref == "source-chunk-builder":
        # This is the formal provenance boundary.  Any further internal LLM
        # context-window splitting remains an execution artifact only.
        chunk_set_id = str(params.get("chunk_set_id") or "")
        if not chunk_set_id:
            raise ValueError("Source Chunk Builder 缺少 chunk_set_id")
        return [{**value, "chunk_set_id": chunk_set_id,
                 "source_chunk_id": hashlib.sha256(f"{chunk_set_id}:{value['chunk_index']}".encode("utf-8")).hexdigest()} for value in values]
    if ref == "faq-table-row-builder":
        documents = [{**value, "table_rows": value.get("_faq_table_rows") or []} for value in values]
        rows = normalize_faq_rows(documents)
        document = documents[0]
        return [{
            "source_id": document["source_id"],
            "source_version_id": document["source_version_id"],
            "filename": document["filename"],
            "content": row["full_text"],
            "chunk_index": index,
            "faq": {key: value for key, value in row.items() if key not in {"row_number", "sheet"}},
            "anchor": {"file": document["filename"], "row": row["row_number"], "sheet": row["sheet"], "chunk_index": index},
        } for index, row in enumerate(rows)]
    if ref == "faq-record-mapper":
        kind = str(params.get("knowledge_type") or "")
        if kind != FAQ_TYPE_CODE:
            raise ValueError("FAQ Record Mapper 只允许 qa-agent-faq")
        contract = type_contracts.get(kind)
        if not contract:
            raise ValueError("qa-agent-faq 缺少已发布契约")
        outcome = (generation if generation is not None else {}).setdefault(kind, {"successful": [], "failed": [], "targeted": []})
        result: list[dict[str, Any]] = []
        for chunk in values:
            scope_key = (kind, str(chunk["source_version_id"]), str(chunk["source_chunk_id"]))
            if retry_scope is not None and scope_key not in retry_scope:
                continue
            outcome["targeted"].append(chunk)
            try:
                data = dict(chunk.get("faq") or {})
                candidate = {
                    "source_knowledge_id": f"faq:{data['aq_id']}",
                    "canonical_content": data["full_text"],
                    "data_json": data,
                    "source_version_ids": [chunk["source_version_id"]],
                    "source_chunk_id": chunk["source_chunk_id"],
                    "source_anchor": f"{chunk['filename']}#row-{(chunk.get('anchor') or {}).get('row')}",
                    "anchor_json": dict(chunk.get("anchor") or {}),
                    "evidence_text": chunk["content"],
                    "is_primary": True,
                }
                result.append(candidate); outcome["successful"].append(chunk)
                if store and job_id:
                    store.record_chunk_generation(job_id, kind, chunk, status="completed", candidate_count=1)
            except Exception as exc:
                outcome["failed"].append({**chunk, "error": str(exc)})
                if store and job_id:
                    store.record_chunk_generation(job_id, kind, chunk, status="failed", error=str(exc))
        return result
    if ref in {"prompt-generator", "qa-generator", "graph-extractor", "structured-knowledge-generator", "multihop-qa"}:
        kind = str(params.get("knowledge_type", ""))
        if ref == "qa-generator": kind = "qa"
        if ref == "graph-extractor": kind = "graph"
        mode = str(params.get("graph_mode") or "")
        output_key = f"graph:{mode}" if kind == "graph" and mode else kind
        if not kind:
            raise ValueError("知识生成节点缺少 knowledge_type")
        contract = type_contracts.get(output_key) or type_contracts.get(kind)
        if not contract and kind != "text":
            raise ValueError(f"知识类型 {kind} 缺少已发布契约")
        llm_serving = str(params.get("llm_serving") or DEFAULT_LLM_SERVING_ID).strip()
        _initialize_llm_servings().require(llm_serving)
        outcome = (generation if generation is not None else {}).setdefault(output_key, {"successful": [], "failed": [], "targeted": []})
        result: list[dict[str, Any]] = []
        for chunk in values:
            scope_key = (output_key, str(chunk["source_version_id"]), str(chunk["source_chunk_id"]))
            if retry_scope is not None and scope_key not in retry_scope:
                continue
            outcome["targeted"].append(chunk)
            try:
                if ref in {"prompt-generator", "structured-knowledge-generator"} and kind == "text" and operator_version == 6:
                    if not contract:
                        raise ValueError("文本生成缺少已发布的文本知识契约")
                    candidates = _generated_text_candidates(
                        sources[chunk["source_id"]], versions[chunk["source_version_id"]], chunk,
                        contract, params, llm_serving,
                    )
                else:
                    candidates = _candidates(
                        sources[chunk["source_id"]], versions[chunk["source_version_id"]], output_key, chunk,
                        contract=contract, llm_serving=llm_serving,
                    )
                outcome["successful"].append(chunk)
                result.extend(candidates)
                if store and job_id:
                    store.record_chunk_generation(job_id, output_key, chunk, status="completed", candidate_count=len(candidates))
            except Exception as exc:
                outcome["failed"].append({**chunk, "error": str(exc)})
                if store and job_id:
                    store.record_chunk_generation(job_id, output_key, chunk, status="failed", error=str(exc))
        return result
    if ref == "entity-extractor":
        output_key = _graph_output_key(params)
        if not params.get("graph_mode"):
            raise ValueError("entity-extractor 缺少 graph_mode")
        llm_serving = str(params.get("llm_serving") or DEFAULT_LLM_SERVING_ID).strip()
        _initialize_llm_servings().require(llm_serving)
        outcome = (generation if generation is not None else {}).setdefault(output_key, {"successful": [], "failed": [], "targeted": []})
        node_config = _graph_config_for_node(cfg, params)
        result: list[dict[str, Any]] = []
        for chunk in values:
            scope_key = (output_key, str(chunk["source_version_id"]), str(chunk["source_chunk_id"]))
            if retry_scope is not None and scope_key not in retry_scope:
                continue
            outcome["targeted"].append(chunk)
            try:
                entities = _extract_entities(chunk, node_config, llm_serving, params)
                result.append({**chunk, "entities": entities})
                outcome["successful"].append(chunk)
                if store and job_id:
                    store.record_chunk_generation(job_id, output_key, chunk, status="completed", candidate_count=len(entities))
            except Exception as exc:
                outcome["failed"].append({**chunk, "error": str(exc)})
                if store and job_id:
                    store.record_chunk_generation(job_id, output_key, chunk, status="failed", error=str(exc))
        return result
    if ref == "relation-extractor":
        output_key = _graph_output_key(params)
        llm_serving = str(params.get("llm_serving") or DEFAULT_LLM_SERVING_ID).strip()
        _initialize_llm_servings().require(llm_serving)
        node_config = _graph_config_for_node(cfg, params, relation=True)
        outcome = (generation if generation is not None else {}).get(output_key)
        result = []
        for record in values:
            try:
                relations = _extract_relations(record, record.get("entities") or [], node_config, llm_serving)
                result.append({**record, "relations": relations})
            except Exception as exc:
                if outcome is not None:
                    outcome["successful"] = [item for item in outcome.get("successful", []) if str(item.get("source_chunk_id")) != str(record.get("source_chunk_id"))]
                    outcome["failed"].append({**record, "error": str(exc)})
                if store and job_id:
                    store.record_chunk_generation(job_id, output_key, record, status="failed", error=str(exc))
        return result
    if ref == "literal-detector":
        return [_annotate_literals(value) for value in values]
    if ref == "entity-normalizer":
        return [_normalize_entities(value, library_id, cfg) for value in values]
    if ref == "triple-builder":
        result = []
        for record in values:
            result.extend(_build_triples(record, cfg, library_id, sources, versions))
        return result
    if ref == "semantic-relation-builder":
        result = []
        for record in values:
            result.extend(_build_semantic_relations(record, cfg, library_id, sources, versions))
        return result
    if ref == "evidence-binder":
        return _bind_evidence(values)
    if ref == "artifact-merge":
        return [dict(value) for value in values]
    if ref == "quality-evaluator":
        rules = dict((params.get("_resolved_quality_profile") or {}).get("rules") or {})
        pass_score, review_score = float(rules.get("pass_score", 0.8)), float(rules.get("review_score", 0.6))
        return [{**value, "quality_score": float(value.get("quality_score", 1.0)),
                 "quality_status": "pass" if float(value.get("quality_score", 1.0)) >= pass_score
                 else "review" if float(value.get("quality_score", 1.0)) >= review_score else "reject"}
                for value in values]
    if ref == "quality-filter":
        rules = dict((params.get("_resolved_quality_profile") or {}).get("rules") or {})
        pass_score, review_score = float(rules.get("pass_score", 0.8)), float(rules.get("review_score", 0.6))
        result = []
        for value in values:
            score = float(value.get("quality_score", 1.0))
            if score < review_score:
                continue
            result.append({**value, "quality_status": "pass" if score >= pass_score else "review"})
        return result
    if ref == "schema-validator":
        if str(params.get("knowledge_type") or "") == "graph":
            graph_mode = str(params.get("graph_mode") or "")
            for value in values:
                _validate_graph_item(dict(value.get("data_json") or {}), cfg, graph_mode)
        return [dict(value) for value in values]
    if ref == "graph-quality-validator":
        graph_mode = str(params.get("graph_mode") or "")
        report = evaluate_graph_quality([dict(value.get("data_json") or {}) for value in values], cfg, graph_mode)
        if report.get("hard_fail"):
            problems = []
            if report.get("literal_as_entity_count"): problems.append("语义图谱出现字面值实体")
            if report.get("invalid_entity_type_count"): problems.append("非法实体类型")
            if report.get("invalid_relation_type_count"): problems.append("非法关系类型")
            if report.get("missing_evidence_count"): problems.append("缺少 Evidence")
            raise ValueError("图谱质量门禁未通过：" + "；".join(problems or ["质量不达标"]))
        return [{**value, "graph_quality": report} for value in values]
    if ref in {"source-binding", "knowledge-diff", "deduplicate", "prompted-refiner"}:
        # Candidate-only operators deliberately never mutate published items.
        if ref == "deduplicate":
            unique: dict[str, dict[str, Any]] = {}
            for value in values:
                unique.setdefault(str(value.get("source_knowledge_id")), value)
            return list(unique.values())
        return [dict(value) for value in values]
    raise ValueError(f"Runner 没有批准的算子 Adapter：{ref}")


def _builtin_dispatch(ref: str, params: dict[str, Any], inputs: list[dict[str, Any]], runtime: dict[str, Any]) -> list[dict[str, Any]]:
    """Bridge the version-pinned OperatorExecutor contract onto the legacy operator table."""
    return _run_operator(
        ref, dict(params or {}), list(inputs),
        root_documents=runtime["root_documents"],
        sources=runtime["sources"],
        versions=runtime["versions"],
        type_contracts=runtime["type_contracts"],
        job_id=runtime.get("job_id"),
        store=runtime.get("store"),
        retry_scope=runtime.get("retry_scope"),
        generation=runtime.get("generation"),
        graph_config=runtime.get("graph_config"),
        sink_libraries=runtime.get("sink_libraries"),
        operator_version=runtime.get("operator_version"),
    )


def execute_source_preparation(store: V7Store, objects, job_id: str, *, lease_owner: str | None = None) -> dict:
    context = store.source_preparation_context(job_id)
    job, version, source = context["job"], context["version"], context["source"]
    chunk_set = context["chunk_set"]
    chunk_params = _operator_params(context.get("definition"), "semantic-chunker")
    if job.status not in {"queued", "running"}:
        return {"id": job.id, "status": job.status, "idempotent": True}
    run = store.start_source_preparation_flow_run(job_id)
    created_parser_keys: list[str] = []
    persisted_parser_keys: set[str] = set()
    try:
        versions = {version.id: version}; sources = {source.id: source}
        documents = _documents_for_versions(objects, [version], sources, run["id"], created_parser_keys)
        store.record_document_irs(run["id"], documents)
        parser_ids = store.record_flow_node(
            run["id"], "parse::parser", [], [{**item, "_artifact_type": "document_ir"} for item in documents],
            operator_code="document-parser",
        )
        persisted_parser_keys.update(created_parser_keys)
        documents = [{key: value for key, value in item.items() if key != "_parser_artifacts"} for item in documents]
        cleaned = documents
        last_ids = parser_ids
        for node_id, ref in (
            ("clean::null", "null-filter"), ("clean::language", "language-filter"),
            ("clean::clean", "text-cleaner"), ("clean::space", "whitespace-cleaner"),
            ("clean::normalize", "text-normalizer"),
        ):
            cleaned = _run_operator(ref, {}, cleaned, root_documents=documents, sources=sources, versions=versions,
                                    type_contracts={})
            last_ids = store.record_flow_node(run["id"], node_id, last_ids, cleaned, operator_code=ref)
        if any(item.get("_faq_table_rows") for item in documents):
            chunk_values = _run_operator("faq-table-row-builder", {}, cleaned, root_documents=documents,
                                         sources=sources, versions=versions, type_contracts={})
            chunk_values = [{**item, "anchor": {**dict(item.get("anchor") or {}), "faq": dict(item.get("faq") or {})}}
                            for item in chunk_values]
            chunk_ids = store.record_flow_node(run["id"], "chunk::faq-rows", last_ids, chunk_values,
                                               operator_code="faq-table-row-builder")
        else:
            chunk_values = _run_operator("semantic-chunker", chunk_params, cleaned, root_documents=documents,
                                         sources=sources, versions=versions, type_contracts={})
            chunk_ids = store.record_flow_node(run["id"], "chunk::chunk", last_ids, chunk_values,
                                               operator_code="semantic-chunker")
        formal = _run_operator("source-chunk-builder", {"chunk_set_id": chunk_set.id}, chunk_values, root_documents=documents,
                               sources=sources, versions=versions, type_contracts={})
        store.record_source_chunks(run["id"], formal)
        store.record_flow_node(run["id"], "chunk::source-chunks", chunk_ids, formal,
                               operator_code="source-chunk-builder")
        store.finish_flow_run(run["id"], status="completed")
        return {**store.finish_source_preparation(job_id), "flow_run_id": run["id"]}
    except Exception as exc:
        store.finish_flow_run(run["id"], str(exc), status="failed")
        result = store.finish_source_preparation(job_id, str(exc))
        for object_key in created_parser_keys:
            if object_key not in persisted_parser_keys:
                try: objects.delete_key(object_key)
                except Exception: logger.exception("清理 Source Preparation Parser Artifact 失败")
        return {**result, "flow_run_id": run["id"]}


def execute_job(store: V7Store, objects, job_id: str, *, lease_owner: str | None = None) -> dict:
    configure_llm_serving_registry(store.sessions, os.getenv("DATAFORGE_CONFIG_ENCRYPTION_KEY"))
    if lease_owner:
        store.assert_work_lease("knowledge", job_id, lease_owner)
    with store.sessions() as session:
        job = session.get(KnowledgeJob, job_id)
        if not job:
            raise ValueError("知识任务不存在")
        if job.status not in {"running", "queued"}:
            return {"id": job.id, "status": job.status, "idempotent": True}
        versions_list = session.scalars(select(SourceVersion).where(SourceVersion.id.in_(job.source_version_ids))).all()
        versions = {version.id: version for version in versions_list}
        sources = {source.id: source for source in session.scalars(select(Source).where(Source.id.in_([version.source_id for version in versions_list])))}
        sink_libraries = dict(job.sink_library_ids or job.output_library_ids)
    retry_scope = store.retry_chunk_scope(job_id)
    flow_run = store.start_flow_run(job_id)
    outputs: dict[str, list[dict[str, Any]]] = {}
    artifact_ids: dict[str, list[str]] = {}
    changes: dict[str, dict[str, int]] = {}
    committed_sinks: set[str] = set()
    sink_errors: dict[str, str] = {}
    node_errors: dict[str, str] = {}
    generation: dict[str, dict[str, list[dict[str, Any]]]] = {}
    current_source_chunks: list[dict[str, Any]] = []
    created_parser_keys: list[str] = []
    persisted_parser_keys: set[str] = set()
    try:
        definition = store.template_definition_for_job(job_id)
        type_contracts = store.type_contracts_for_job(job_id)
        graph_config = _graph_config_from_contracts(type_contracts)
        incoming = _incoming(definition)
        root_documents = store.reviewed_chunks_for_job(job_id)
        current_source_chunks = [dict(value) for value in root_documents]
        catalog = catalog_by_code()
        registry = build_builtin_registry(_builtin_dispatch, catalog)
        runtime = {
            "root_documents": root_documents, "sources": sources, "versions": versions,
            "type_contracts": type_contracts, "job_id": job_id, "store": store,
            "retry_scope": retry_scope, "generation": generation,
            "graph_config": graph_config, "sink_libraries": sink_libraries,
        }
        sink_nodes: list[dict[str, Any]] = []
        for node in definition.get("nodes", []):
            if store.is_job_cancelled(job_id):
                store.finish_flow_run(flow_run["id"], "cancelled")
                return {"id": job_id, "status": "cancelled", "flow_run_id": flow_run["id"]}
            node_id = node["id"]
            source_nodes = incoming.get(node_id, [])
            input_values = [value for source_id in source_nodes for value in outputs.get(source_id, [])]
            input_ids = [artifact_id for source_id in source_nodes for artifact_id in artifact_ids.get(source_id, [])]
            if node.get("kind") == "knowledge_sink":
                sink_nodes.append(node)
                continue
            ref = str(node.get("ref"))
            operator_version = int(node.get("operator_version") or catalog.get(ref, {}).get("version", 1))
            resolved_params = dict(node.get("params") or {})
            failed_upstream = [source_id for source_id in source_nodes if source_id in node_errors]
            if failed_upstream:
                message = "上游节点失败，已跳过：" + "、".join(failed_upstream)
                node_errors[node_id] = message; outputs[node_id] = []
                artifact_ids[node_id] = store.record_flow_node(flow_run["id"], node_id, input_ids, [], error=message, operator_code=ref, operator_version=operator_version, resolved_parameters=resolved_params)
                continue
            try:
                executor = registry.resolve(ref, operator_version)
                result = executor.execute(
                    inputs=input_values, params=resolved_params,
                    context=OperatorExecutionContext(flow_run_id=flow_run["id"], node_id=node_id, runtime=runtime),
                )
                values = result.outputs
            except Exception as exc:
                node_errors[node_id] = str(exc); outputs[node_id] = []
                artifact_ids[node_id] = store.record_flow_node(flow_run["id"], node_id, input_ids, [], error=str(exc), operator_code=ref, operator_version=operator_version, resolved_parameters=resolved_params)
                continue
            if ref == "document-parser":
                store.record_document_irs(flow_run["id"], values)
            elif ref == "source-chunk-builder":
                store.record_source_chunks(flow_run["id"], values)
                current_source_chunks = [dict(value) for value in values]
            outputs[node_id] = values
            artifact_type = "execution"
            if values:
                values = [{**value, "_artifact_type": artifact_type} for value in values]
                outputs[node_id] = values
            artifact_values = values
            if ref == "document-parser":
                artifact_values = [{key: value for key, value in item.items() if key != "_faq_table_rows"} for item in values]
            artifact_ids[node_id] = store.record_flow_node(flow_run["id"], node_id, input_ids, artifact_values, operator_code=ref, operator_version=operator_version, resolved_parameters=resolved_params)
            if ref == "document-parser":
                persisted_parser_keys.update(created_parser_keys)
                outputs[node_id] = [{key: value for key, value in item.items() if key != "_parser_artifacts"} for item in outputs[node_id]]
        # All generation gates finish before any Knowledge Sink writes.  A
        # failed chunk stays out of a Sink's replacement range, so successful
        # neighbouring chunks can publish without erasing it.
        for node in sink_nodes:
            if lease_owner:
                store.assert_work_lease("knowledge", job_id, lease_owner)
            node_id = node["id"]
            source_nodes = incoming.get(node_id, [])
            input_values = [value for source_id in source_nodes for value in outputs.get(source_id, [])]
            input_ids = [artifact_id for source_id in source_nodes for artifact_id in artifact_ids.get(source_id, [])]
            knowledge_type = node["knowledge_type"]
            output_key = str(node.get("output_key") or (f"graph:{node.get('graph_mode')}" if knowledge_type == "graph" and node.get("graph_mode") else knowledge_type))
            try:
                failed_upstream = [source_id for source_id in source_nodes if source_id in node_errors]
                if failed_upstream:
                    raise ValueError("上游节点失败，Sink 已跳过：" + "、".join(failed_upstream))
                outcomes = generation.get(output_key, {"successful": [], "failed": [], "targeted": []})
                successful = outcomes["successful"]
                if successful:
                    changes[output_key] = store.apply_knowledge_output(
                        job_id, output_key, input_values, successful_chunks=successful,
                    )
                    committed_sinks.add(output_key)
                else:
                    changes[output_key] = {"ADD": 0, "UPDATE": 0, "INACTIVE": 0, "UNCHANGED": 0}
                outputs[node_id] = [{"_artifact_type": f"knowledge_item:{output_key}", "knowledge_type": knowledge_type, "graph_mode": node.get("graph_mode"), "change": changes[output_key]}]
                artifact_ids[node_id] = store.record_flow_node(flow_run["id"], node_id, input_ids, outputs[node_id])
            except Exception as exc:
                sink_errors[output_key] = str(exc)
                outputs[node_id] = []
                artifact_ids[node_id] = store.record_flow_node(flow_run["id"], node_id, input_ids, [], error=str(exc))
        missing_generation = set(sink_libraries) - set(generation)
        for knowledge_type in missing_generation:
            sink_errors[knowledge_type] = "流程没有执行该知识类型的生成节点"
        successful_chunks = [entry for outcome in generation.values() for entry in outcome["successful"]]
        if not successful_chunks or not committed_sinks:
            detail = "; ".join(f"{kind}: {error}" for kind, error in sink_errors.items())
            if not detail:
                detail = "所有知识生成分块均失败" if not successful_chunks else "所有 Knowledge Sink 均未成功提交"
            store.finish_flow_run(flow_run["id"], detail)
            if lease_owner:
                store.assert_work_lease("knowledge", job_id, lease_owner)
            store.mark_job_failed(job_id, detail)
            return {"id": job_id, "status": "failed", "flow_run_id": flow_run["id"], "changes": changes, "sink_errors": sink_errors}
        warnings = [{"knowledge_type": kind, "source_version_id": item["source_version_id"], "source_chunk_id": item["source_chunk_id"], "chunk_index": item["chunk_index"], "error": item["error"]} for kind, outcome in generation.items() for item in outcome["failed"]]
        warnings.extend({"node_id": node_id, "error": message} for node_id, message in node_errors.items())
        warnings.extend({"output_key": key, "error": message} for key, message in sink_errors.items())
        # Only after every sink has applied its successful chunk range may an
        # older missing chunk be withdrawn.  A failure in any target type keeps
        # the whole source-version history, including on a failed-only retry.
        if not warnings:
            if lease_owner:
                store.assert_work_lease("knowledge", job_id, lease_owner)
            cleanup_source_versions = store.completed_source_versions_for_cleanup(
                job_id, set(sink_libraries), current_source_chunks,
            )
            if cleanup_source_versions:
                for knowledge_type in sink_libraries:
                    store.apply_knowledge_output(
                        job_id, knowledge_type, [], successful_chunks=current_source_chunks,
                        replace_absent_source_versions=cleanup_source_versions,
                        cleanup_only=True,
                    )
        store.finish_flow_run(flow_run["id"], "部分分块生成失败" if warnings else None,
                              status="completed_with_warnings" if warnings else "completed")
        if lease_owner:
            store.assert_work_lease("knowledge", job_id, lease_owner)
        store.complete_job(job_id, warnings=warnings)
        vector_jobs = {
            library_id: store.create_vector_sync_jobs(library_id)
            for output_key, library_id in sink_libraries.items()
            if output_key in committed_sinks
        }
        return {"id": job_id, "status": "completed_with_warnings" if warnings else "completed", "flow_run_id": flow_run["id"], "changes": changes, "warnings": warnings, "vector_sync_jobs": vector_jobs}
    except Exception as exc:
        if lease_owner:
            try:
                store.assert_work_lease("knowledge", job_id, lease_owner)
            except ValueError:
                store.finish_flow_run(flow_run["id"], "任务执行租约已失效")
                raise
        for object_key in created_parser_keys:
            if object_key not in persisted_parser_keys:
                try:
                    objects.delete_key(object_key)
                except Exception:
                    pass
        store.mark_source_versions_failed([item.id for item in versions_list], str(exc))
        store.finish_flow_run(flow_run["id"], str(exc))
        store.mark_job_failed(job_id, str(exc))
        raise


def _reachable_nodes(definition: dict[str, Any], start_node_id: str, mode: str) -> set[str]:
    if mode == "node_only": return {start_node_id}
    outgoing: dict[str, list[str]] = {str(node["id"]): [] for node in definition.get("nodes", [])}
    for edge in definition.get("edges", []): outgoing.setdefault(str(edge["source"]), []).append(str(edge["target"]))
    result, queue = set(), [start_node_id]
    while queue:
        current = queue.pop(0)
        if current in result: continue
        result.add(current); queue.extend(outgoing.get(current, []))
    return result


def _candidate_chunks(values: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: dict[tuple[str, str], dict[str, Any]] = {}
    for value in values:
        chunk_id = str(value.get("source_chunk_id") or "")
        for version_id in value.get("source_version_ids") or []:
            if chunk_id:
                result[(str(version_id), chunk_id)] = {"source_version_id": str(version_id), "source_chunk_id": chunk_id,
                                                        "chunk_index": int((value.get("anchor_json") or {}).get("chunk_index", 0))}
    return list(result.values())


def execute_debug_run(store: V7Store, objects, flow_run_id: str) -> dict[str, Any]:
    """Execute a full or derived Debug Run without mutating formal knowledge."""
    context = store.debug_run_context(flow_run_id)
    definition = context["definition"]
    incoming = _incoming(definition)
    selected = ({str(node["id"]) for node in definition.get("nodes", [])}
                if context["mode"] == "debug_full"
                else _reachable_nodes(definition, context["start_node_id"], context["mode"]))
    versions = {item.id: item for item in context["versions"]}
    sources = context["sources"]
    root_documents = context["root_documents"]
    type_contracts = context["type_contracts"]
    graph_config = _graph_config_from_contracts(type_contracts)
    catalog = catalog_by_code()
    registry = build_builtin_registry(_builtin_dispatch, catalog)
    outputs: dict[str, list[dict[str, Any]]] = {}
    artifact_ids: dict[str, list[str]] = {}
    failed: dict[str, str] = {}
    generation: dict[str, dict[str, list[dict[str, Any]]]] = {}
    previews: list[dict[str, Any]] = []
    runtime = {
        "root_documents": root_documents, "sources": sources, "versions": versions,
        "type_contracts": type_contracts, "job_id": None, "store": None,
        "retry_scope": None, "generation": generation, "graph_config": graph_config,
        "sink_libraries": context["sink_libraries"],
    }
    try:
        for node in definition.get("nodes", []):
            node_id = str(node["id"])
            if node_id not in selected:
                continue
            if store.is_flow_run_cancelled(flow_run_id):
                store.finish_flow_run(flow_run_id, "已协作停止", status="cancelled")
                return {"id": flow_run_id, "status": "cancelled"}
            source_nodes = incoming.get(node_id, [])
            input_values: list[dict[str, Any]] = []
            input_ids: list[str] = []
            for source_id in source_nodes:
                if source_id in selected:
                    source = {"values": outputs.get(source_id, []), "ids": artifact_ids.get(source_id, [])}
                else:
                    source = context["parent_outputs"].get(source_id, {"values": [], "ids": []})
                input_values.extend(source["values"]); input_ids.extend(source["ids"])
            failed_upstream = [source_id for source_id in source_nodes if source_id in selected and source_id in failed]
            if failed_upstream:
                message = "上游节点失败，已跳过：" + "、".join(failed_upstream)
                failed[node_id] = message; outputs[node_id] = []
                artifact_ids[node_id] = store.record_flow_node(
                    flow_run_id, node_id, input_ids, [], error=message, status="skipped",
                )
                continue
            if node.get("kind") == "knowledge_sink":
                output_key = str(node.get("output_key") or (
                    f"graph:{node.get('graph_mode')}" if node.get("knowledge_type") == "graph" and node.get("graph_mode")
                    else node.get("knowledge_type")
                ))
                target = dict((context.get("sink_preview_targets") or {}).get(output_key) or {})
                library_id = target.get("knowledge_library_id") or context["sink_libraries"].get(output_key)
                baseline_kind = str(target.get("baseline_kind") or "knowledge_library")
                if baseline_kind == "knowledge_library" and not library_id:
                    message = "Sink 缺少预览目标知识库"; failed[node_id] = message
                    artifact_ids[node_id] = store.record_flow_node(flow_run_id, node_id, input_ids, [], error=message)
                    continue
                successful = generation.get(output_key, {}).get("successful") or _candidate_chunks(input_values)
                preview = store.stage_sink_preview(
                    flow_run_id, output_key, library_id, input_values, successful,
                    quality={"candidate_count": len(input_values), "status": "preview_only"},
                    baseline_kind=baseline_kind,
                )
                previews.append(preview)
                outputs[node_id] = [{"_artifact_type": f"knowledge_preview:{output_key}", **preview}]
                artifact_ids[node_id] = store.record_flow_node(
                    flow_run_id, node_id, input_ids, outputs[node_id], status="preview_ready",
                )
                continue
            ref = str(node.get("ref"))
            params = {**dict(node.get("params") or {}), **dict((context["parameter_overrides"] or {}).get(node_id) or {})}
            params.pop("force_ocr", None)
            operator_version = int(node.get("operator_version") or (catalog.get(ref) or {}).get("version", 1))
            try:
                result = registry.resolve(ref, operator_version).execute(
                    inputs=input_values, params=params,
                    context=OperatorExecutionContext(flow_run_id=flow_run_id, node_id=node_id, runtime=runtime),
                )
                outputs[node_id] = result.outputs
                item = catalog.get(ref) or {}
                recorded = [{**value, "_artifact_type": item.get("output", "execution")} for value in result.outputs]
                artifact_ids[node_id] = store.record_flow_node(
                    flow_run_id, node_id, input_ids, recorded, operator_code=ref,
                    operator_version=operator_version, resolved_parameters=params,
                    metrics={"input_records": len(input_values), "output_records": len(result.outputs)},
                )
            except Exception as exc:
                message = str(exc); failed[node_id] = message; outputs[node_id] = []
                artifact_ids[node_id] = store.record_flow_node(
                    flow_run_id, node_id, input_ids, [], error=message, operator_code=ref,
                    operator_version=operator_version, resolved_parameters=params,
                )
        if failed:
            status = "completed_with_warnings" if previews else "failed"
            detail = "；".join(f"{node_id}: {message}" for node_id, message in failed.items())
            store.finish_flow_run(flow_run_id, detail, status=status)
            return {"id": flow_run_id, "status": status, "previews": previews, "errors": failed}
        store.finish_flow_run(flow_run_id, status="completed")
        return {"id": flow_run_id, "status": "completed", "previews": previews}
    except Exception as exc:
        store.finish_flow_run(flow_run_id, str(exc), status="failed")
        raise


def execute_flow_run(store: V7Store, objects, flow_run_id: str) -> dict[str, Any]:
    if store.flow_run_is_debug(flow_run_id):
        return execute_debug_run(store, objects, flow_run_id)
    return execute_derived_run(store, objects, flow_run_id)


def execute_derived_run(store: V7Store, objects, flow_run_id: str) -> dict[str, Any]:
    context = store.derived_run_context(flow_run_id); definition = context["definition"]; incoming = _incoming(definition)
    selected = _reachable_nodes(definition, context["start_node_id"], context["mode"])
    by_id = {str(node["id"]): node for node in definition.get("nodes", [])}
    versions_list = context["versions"]; versions = {item.id: item for item in versions_list}; sources = context["sources"]
    type_contracts = store.type_contracts_for_job(context["job_id"]); catalog = catalog_by_code(); parent_outputs = context["parent_outputs"]
    graph_config = _graph_config_from_contracts(type_contracts)
    registry = build_builtin_registry(_builtin_dispatch, catalog)
    outputs: dict[str, list[dict[str, Any]]] = {}; artifact_ids: dict[str, list[str]] = {}; failed: set[str] = set()
    generation: dict[str, dict[str, list[dict[str, Any]]]] = {}; previews = []; created_parser_keys: list[str] = []
    try:
        start_node = by_id[context["start_node_id"]]
        override = dict((context["parameter_overrides"] or {}).get(context["start_node_id"]) or {})
        root_documents = _documents_for_versions(objects, versions_list, sources, flow_run_id, created_parser_keys,
                                                 force_ocr=bool(override.get("force_ocr"))) if start_node.get("ref") == "document-parser" else []
        for node in definition.get("nodes", []):
            node_id = str(node["id"])
            if node_id not in selected: continue
            if store.is_flow_run_cancelled(flow_run_id):
                store.finish_flow_run(flow_run_id, "已协作停止", status="cancelled"); return {"id": flow_run_id, "status": "cancelled"}
            source_nodes = incoming.get(node_id, []); input_values: list[dict[str, Any]] = []; input_ids: list[str] = []
            for source_id in source_nodes:
                source = {"values": outputs.get(source_id, []), "ids": artifact_ids.get(source_id, [])} if source_id in selected else parent_outputs.get(source_id, {"values": [], "ids": []})
                input_values.extend(source["values"]); input_ids.extend(source["ids"])
            failed_upstream = [source_id for source_id in source_nodes if source_id in selected and source_id in failed]
            if failed_upstream:
                message = "上游节点失败，已跳过：" + "、".join(failed_upstream); failed.add(node_id); outputs[node_id] = []
                artifact_ids[node_id] = store.record_flow_node(flow_run_id, node_id, input_ids, [], error=message, status="skipped"); continue
            if node.get("kind") == "knowledge_sink":
                output_key = str(node.get("output_key") or node.get("knowledge_type")); library_id = context["sink_libraries"].get(output_key)
                if not library_id:
                    failed.add(node_id); artifact_ids[node_id] = store.record_flow_node(flow_run_id, node_id, input_ids, [], error="Sink 缺少目标知识库"); continue
                successful = generation.get(output_key, {}).get("successful") or _candidate_chunks(input_values)
                preview = store.stage_sink_preview(flow_run_id, output_key, library_id, input_values, successful); previews.append(preview)
                outputs[node_id] = [{"_artifact_type": f"knowledge_preview:{output_key}", **preview}]
                artifact_ids[node_id] = store.record_flow_node(flow_run_id, node_id, input_ids, outputs[node_id], status="awaiting_commit"); continue
            ref = str(node.get("ref")); params = {**dict(node.get("params") or {}), **dict((context["parameter_overrides"] or {}).get(node_id) or {})}; params.pop("force_ocr", None)
            try:
                operator_version = int(node.get("operator_version") or (catalog.get(ref) or {}).get("version", 1))
                values = registry.resolve(ref, operator_version).execute(
                    inputs=input_values, params=params,
                    context=OperatorExecutionContext(flow_run_id=flow_run_id, node_id=node_id, runtime={
                        "root_documents": root_documents, "sources": sources, "versions": versions,
                        "type_contracts": type_contracts, "generation": generation,
                        "graph_config": graph_config, "sink_libraries": context.get("sink_libraries"),
                    }),
                ).outputs
                outputs[node_id] = values; item = catalog.get(ref) or {}
                recorded = [{**value, "_artifact_type": item.get("output", "execution")} for value in values]
                artifact_ids[node_id] = store.record_flow_node(flow_run_id, node_id, input_ids, recorded, operator_code=ref,
                                                               operator_version=int(node.get("operator_version") or item.get("version", 1)), resolved_parameters=params,
                                                               metrics={"input_records": len(input_values), "output_records": len(values)})
            except Exception as exc:
                failed.add(node_id); outputs[node_id] = []
                artifact_ids[node_id] = store.record_flow_node(flow_run_id, node_id, input_ids, [], error=str(exc), operator_code=ref,
                                                               operator_version=int(node.get("operator_version") or (catalog.get(ref) or {}).get("version", 1)), resolved_parameters=params)
        if previews:
            store.finish_flow_run(flow_run_id, status="awaiting_commit"); return {"id": flow_run_id, "status": "awaiting_commit", "previews": previews}
        status = "failed" if context["start_node_id"] in failed else "completed"
        store.finish_flow_run(flow_run_id, "派生节点执行失败" if status == "failed" else None, status=status); return {"id": flow_run_id, "status": status}
    except Exception as exc:
        store.finish_flow_run(flow_run_id, str(exc), status="failed"); raise


def create_app(settings: Settings | None = None, *, check_schema: bool = True) -> FastAPI:
    resolved = settings or Settings.load(); resolved.ensure_directories()
    store = V7Store(resolved.platform_database_url)
    if check_schema: store.assert_schema_current()
    configure_llm_serving_registry(store.sessions, resolved.config_encryption_key)
    _initialize_llm_servings()
    objects = _objects(resolved); app = FastAPI(title="DataForge V7 Runner", version="7.0.0")
    active_requests: set[str] = set()
    active_lock = threading.RLock()
    heartbeat_stop = threading.Event()
    instance_id = f"runner:{os.getenv('HOSTNAME') or 'local'}:{os.getpid()}"

    def heartbeat_loop() -> None:
        while not heartbeat_stop.is_set():
            with active_lock:
                current = sorted(active_requests)
            try:
                store.upsert_component_heartbeat(
                    "runner", instance_id, version="7.0.0", current_job_id=current[0] if current else None,
                    details={"active_request_ids": current, "active_count": len(current)},
                )
            except Exception:
                logger.exception("Runner 心跳写入失败")
            heartbeat_stop.wait(15.0)

    @app.on_event("startup")
    def start_heartbeat() -> None:
        heartbeat_stop.clear()
        threading.Thread(target=heartbeat_loop, name="runner-heartbeat", daemon=True).start()

    @app.on_event("shutdown")
    def stop_heartbeat() -> None:
        heartbeat_stop.set()

    def verify(authorization: str | None) -> None:
        expected = f"Bearer {resolved.runner_service_token}" if resolved.runner_service_token else None
        if expected and authorization != expected: raise HTTPException(status_code=403, detail="Runner 服务凭据无效")

    @app.post("/internal/jobs", status_code=202)
    def run(payload: RunRequest, authorization: str | None = Header(None)):
        verify(authorization)
        request_id = payload.job_id or payload.source_preparation_job_id or payload.flow_run_id or "invalid"
        with active_lock:
            active_requests.add(request_id)
        try:
            if payload.flow_run_id: return execute_flow_run(store, objects, payload.flow_run_id)
            if payload.source_preparation_job_id:
                if not payload.lease_owner:
                    raise ValueError("Source Preparation 请求缺少 lease_owner")
                return execute_source_preparation(
                    store, objects, payload.source_preparation_job_id, lease_owner=payload.lease_owner,
                )
            if payload.job_id:
                if not payload.lease_owner:
                    raise ValueError("知识任务请求缺少 lease_owner")
                return execute_job(store, objects, payload.job_id, lease_owner=payload.lease_owner)
            raise ValueError("job_id、source_preparation_job_id 或 flow_run_id 至少提供一个")
        except ValueError as exc: raise HTTPException(status_code=422, detail=str(exc)) from exc
        finally:
            with active_lock:
                active_requests.discard(request_id)

    @app.get("/internal/runtime")
    def runtime_status(authorization: str | None = Header(None)):
        verify(authorization)
        return runtime

    @app.get("/internal/health")
    def health(authorization: str | None = Header(None)):
        verify(authorization)
        with store.engine.connect() as connection:
            connection.exec_driver_sql("SELECT 1")
        with active_lock:
            current = sorted(active_requests)
        registry = _initialize_llm_servings()
        return {"status": "healthy", "version": "7.0.0", "active_request_ids": current,
                "serving_ids": sorted(registry.serving_ids)}

    @app.post("/internal/diagnostics/{component}")
    def diagnostics(component: str, authorization: str | None = Header(None)):
        verify(authorization)
        if component == "runner":
            return health(authorization)
        if component == "llm":
            started = time.monotonic()
            value = _llm_json('只返回 {"ok": true}', system="你是健康检查器，只返回 JSON。")
            if value.get("ok") is not True:
                raise HTTPException(status_code=503, detail="LLM 探针响应无效")
            return {"status": "healthy", "serving_id": DEFAULT_LLM_SERVING_ID,
                    "latency_ms": round((time.monotonic() - started) * 1000)}
        if component == "mineru":
            base_url = os.getenv("DATAFORGE_MINERU_URL", "http://mineru-api:8000").rstrip("/")
            response = httpx.get(f"{base_url}/health", timeout=5.0)
            response.raise_for_status()
            health_payload = dict(response.json())
            # One-page, in-memory PDF; no Source, Artifact or object-store write is created.
            parsed = parse_with_mineru(filename="dataforge-health.pdf", payload=_diagnostic_pdf())
            return {"status": "healthy", "version": parsed.version, "backend": parsed.backend,
                    "health": health_payload, "model_inference_verified": True}
        raise HTTPException(status_code=404, detail="不支持的 Runner 诊断组件")
    return app


app = FastAPI(title="DataForge V7 Runner bootstrap")


def main() -> None:
    import uvicorn
    uvicorn.run(create_app(check_schema=True), host="0.0.0.0", port=int(os.getenv("DATAFORGE_RUNNER_PORT", "8010")))

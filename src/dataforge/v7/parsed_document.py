"""Authoritative ParsedDocument content and object-store helpers."""
from __future__ import annotations

import csv
from copy import deepcopy
import hashlib
import io
import json
from html.parser import HTMLParser
from pathlib import Path
from typing import Any


PARSER_REVISION = "parsed-document-v1"
TABLE_FORMAT = "dataforge.table-v1"


def canonical_json(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def object_key(parsed_document_id: str, name: str) -> str:
    return f"parsed-documents/{parsed_document_id}/{name}"


def persist_content(objects, parsed_document_id: str, *, kind: str, content: str | dict,
                    anchors: dict) -> dict[str, Any]:
    if kind == "textual":
        content_format = "markdown"
        content_bytes = str(content).encode("utf-8")
        content_type = "text/markdown; charset=utf-8"
        name = "content.md"
    elif kind == "tabular":
        content_format = "table-v1"
        content_bytes = canonical_json(content)
        content_type = "application/vnd.dataforge.table+json"
        name = "table.json"
    else:
        raise ValueError(f"未知 ParsedDocument kind：{kind}")
    anchor_bytes = canonical_json(anchors)
    stored_content = objects.put_bytes(object_key(parsed_document_id, name), content_bytes, content_type)
    stored_anchors = objects.put_bytes(
        object_key(parsed_document_id, "anchors.json"), anchor_bytes, "application/json"
    )
    return {
        "kind": kind,
        "content_format": content_format,
        "content_ref": f"object:///{stored_content.key}",
        "content_digest": sha256(content_bytes),
        "anchor_map_ref": f"object:///{stored_anchors.key}",
        "anchor_map_digest": sha256(anchor_bytes),
        "object_keys": [stored_content.key, stored_anchors.key],
    }


def load_verified(objects, ref: str, expected_digest: str) -> bytes:
    payload = (objects.get_bytes(ref.removeprefix("object:///"))
               if str(ref).startswith("object:///") else objects.get_blob(ref))
    if sha256(payload) != expected_digest:
        raise ValueError("PARSED_DOCUMENT_DIGEST_MISMATCH")
    return payload


def read_content(objects, parsed_document) -> str | dict:
    payload = load_verified(objects, parsed_document.content_ref, parsed_document.content_digest)
    if parsed_document.kind == "textual":
        return payload.decode("utf-8")
    value = json.loads(payload)
    if not isinstance(value, dict) or value.get("schema") != TABLE_FORMAT:
        raise ValueError("PARSED_DOCUMENT_TABLE_INVALID")
    return value


def read_anchors(objects, parsed_document) -> dict:
    payload = load_verified(objects, parsed_document.anchor_map_ref, parsed_document.anchor_map_digest)
    value = json.loads(payload)
    if not isinstance(value, dict):
        raise ValueError("PARSED_DOCUMENT_ANCHOR_MAP_INVALID")
    return value


def reviewed_text_anchors(anchors: dict, *, content_changed: bool,
                          parent_parsed_document_id: str) -> dict:
    """Preserve source provenance while refusing exact ranges after free-form Markdown edits."""
    value = deepcopy(anchors)
    value["reviewed_from_parsed_document_id"] = parent_parsed_document_id
    value["review_adjusted"] = bool(content_changed)
    if not content_changed:
        return value
    blocks = value.get("blocks")
    if isinstance(blocks, list):
        for block in blocks:
            if not isinstance(block, dict):
                continue
            block["review_adjusted"] = True
            block["precision"] = "parent" if (
                block.get("positions") or block.get("page") or block.get("bbox")
            ) else "unavailable"
            block.pop("char_start", None)
            block.pop("char_end", None)
            block.pop("markdown_range", None)
    value["precision"] = "parent" if blocks else "unavailable"
    return value


def apply_table_cell_updates(table: dict, anchors: dict, updates: list[dict], *,
                             parent_parsed_document_id: str) -> tuple[dict, dict]:
    """Apply review edits without allowing table coordinates or dimensions to drift."""
    content = deepcopy(table)
    reviewed_anchors = deepcopy(anchors)
    if content.get("schema") != TABLE_FORMAT:
        raise ValueError("PARSED_DOCUMENT_TABLE_INVALID")
    sheets = list(content.get("sheets") or [])
    by_coordinate: dict[tuple[int, int, int], dict] = {}
    for sheet in sheets:
        sheet_index = int(sheet.get("sheet_index", -1))
        for row in sheet.get("rows") or []:
            row_index = int(row.get("row_index", -1))
            for cell in row.get("cells") or []:
                by_coordinate[(sheet_index, row_index, int(cell.get("column_index", -1)))] = cell
    seen: set[tuple[int, int, int]] = set()
    for raw in updates:
        coordinate = (
            int(raw.get("sheet_index", -1)), int(raw.get("row_index", -1)),
            int(raw.get("column_index", -1)),
        )
        if coordinate in seen:
            raise ValueError("PARSED_DOCUMENT_CELL_UPDATE_DUPLICATE")
        seen.add(coordinate)
        cell = by_coordinate.get(coordinate)
        if cell is None:
            raise ValueError("PARSED_DOCUMENT_CELL_NOT_FOUND")
        value_type = str(raw.get("value_type") or "")
        value = raw.get("value")
        if value_type == "empty":
            value = None
        elif value_type == "string":
            value = "" if value is None else str(value)
        elif value_type == "number":
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                raise ValueError("PARSED_DOCUMENT_CELL_NUMBER_INVALID")
        elif value_type == "boolean":
            if not isinstance(value, bool):
                raise ValueError("PARSED_DOCUMENT_CELL_BOOLEAN_INVALID")
        else:
            raise ValueError("PARSED_DOCUMENT_CELL_TYPE_INVALID")
        cell["value"], cell["value_type"], cell["review_adjusted"] = value, value_type, True
    reviewed_anchors["reviewed_from_parsed_document_id"] = parent_parsed_document_id
    reviewed_anchors["review_adjusted"] = bool(updates)
    adjusted = set(seen)
    for cell_anchor in reviewed_anchors.get("cells") or []:
        if not isinstance(cell_anchor, dict):
            continue
        coordinate = (
            int(cell_anchor.get("sheet_index", -1)), int(cell_anchor.get("row_index", -1)),
            int(cell_anchor.get("column_index", -1)),
        )
        if coordinate in adjusted:
            cell_anchor["review_adjusted"] = True
    return content, reviewed_anchors


def _cell(column_index: int, value: Any) -> dict[str, Any]:
    if value is None:
        value_type, normalized = "empty", None
    elif isinstance(value, bool):
        value_type, normalized = "boolean", value
    elif isinstance(value, (int, float)):
        value_type, normalized = "number", value
    else:
        value_type, normalized = "string", str(value)
    return {"column_index": column_index, "value": normalized, "value_type": value_type}


def table_v1(filename: str, payload: bytes) -> tuple[dict, dict]:
    suffix = Path(filename).suffix.lower()
    sheets: list[dict[str, Any]] = []
    if suffix == ".csv":
        text = payload.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text, newline="")))
        width = max((len(row) for row in rows), default=0)
        sheets.append({
            "sheet_index": 0, "name": Path(filename).stem or "CSV", "row_count": len(rows),
            "column_count": width,
            "rows": [{"row_index": index, "cells": [_cell(column, row[column] if column < len(row) else None)
                                                        for column in range(width)]}
                     for index, row in enumerate(rows)],
        })
    elif suffix == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        for sheet_index, sheet in enumerate(workbook.worksheets):
            width = int(sheet.max_column or 0)
            height = int(sheet.max_row or 0)
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=height, max_col=width,
                                                             values_only=True)):
                rows.append({"row_index": row_index,
                             "cells": [_cell(column, row[column] if column < len(row) else None)
                                       for column in range(width)]})
            sheets.append({"sheet_index": sheet_index, "name": sheet.title, "row_count": height,
                           "column_count": width, "rows": rows})
    else:
        raise ValueError("table-v1 仅支持 CSV/XLSX")
    content = {"schema": TABLE_FORMAT, "filename": filename, "sheets": sheets}
    anchors = {
        "anchor_version": 2, "source_type": suffix.removeprefix("."),
        "cells": [{"anchor_id": f"sheet:{sheet['sheet_index']}:row:{row['row_index']}:col:{cell['column_index']}",
                   "sheet_index": sheet["sheet_index"], "sheet": sheet["name"],
                   "row_index": row["row_index"], "column_index": cell["column_index"]}
                  for sheet in sheets for row in sheet["rows"] for cell in row["cells"]],
    }
    return content, anchors


class _MarkdownHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self.links: list[str | None] = []

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            self.parts.append("\n\n" + "#" * int(tag[1]) + " ")
        elif tag in {"p", "div", "section", "article", "br", "tr"}:
            self.parts.append("\n")
        elif tag == "li":
            self.parts.append("\n- ")
        elif tag in {"strong", "b"}:
            self.parts.append("**")
        elif tag in {"em", "i"}:
            self.parts.append("*")
        elif tag == "a":
            self.parts.append("[")
            self.links.append(attrs.get("href"))
        elif tag in {"td", "th"}:
            self.parts.append(" | ")

    def handle_endtag(self, tag):
        if tag in {"strong", "b"}:
            self.parts.append("**")
        elif tag in {"em", "i"}:
            self.parts.append("*")
        elif tag == "a":
            href = self.links.pop() if self.links else None
            self.parts.append(f"]({href})" if href else "]")
        elif tag in {"p", "div", "section", "article", "h1", "h2", "h3", "h4", "h5", "h6"}:
            self.parts.append("\n")

    def handle_data(self, data):
        self.parts.append(data)


def html_to_markdown(payload: bytes) -> str:
    parser = _MarkdownHTMLParser()
    parser.feed(payload.decode("utf-8-sig"))
    lines = [line.rstrip() for line in "".join(parser.parts).splitlines()]
    result: list[str] = []
    for line in lines:
        if line.strip() or result and result[-1] != "":
            result.append(line.strip())
    return "\n".join(result).strip() + "\n"


def markdown_from_blocks(blocks: list[dict[str, Any]], fallback: str = "") -> str:
    parts: list[str] = []
    for block in blocks:
        text = str(block.get("text") or "").strip()
        if not text:
            continue
        level = int(block.get("heading_level") or 0)
        parts.append(("#" * min(level, 6) + " " if level else "") + text)
    return "\n\n".join(parts) if parts else str(fallback)
